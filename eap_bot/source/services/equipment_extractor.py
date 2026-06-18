import csv
import io
import json
import logging
import re
from concurrent.futures import ThreadPoolExecutor
from pathlib import Path
from typing import Union

import pdfplumber
import tiktoken

from source.config import settings
from source.schemas.secsgem import (
    Alarm,
    DataVariable,
    EquipmentSpec,
    Event,
    RemoteCommand,
    RCMDParameter,
    State,
    StateTransition,
    StatusVariable,
)
from source.utils.llm_factory import LLMStrategy

logger = logging.getLogger(__name__)

# An SML/SECS-II message code block begins with a bracketed stream/function
# header like "<S2F41 W". The leading "<" distinguishes real SML samples from the
# many bare "S2F41 ..." mentions that appear in surrounding prose and tables.
_SML_HEADER_RE = re.compile(r"^<\s*S\d+F\d+", re.IGNORECASE)


PREDEFINED_QUESTIONS = [
    "What is the tool type or model of this equipment?",
    "What protocol version of SECS/GEM does this equipment support?",
    "Are there any status variables (SVs) defined for monitoring process parameters?",
    "What are the primary data variables (DVs) related to processing?",
    "List any critical events (CEIDs) triggered during a lot start or lot end.",
    "Which alarms are defined for safety or equipment fault conditions?",
    "What remote commands (RCMDs) are supported to control the equipment?",
    "Are there any states or state transition definitions mentioned?",
    "What is the default communication state on power-up?",
    "Describe the process state machine transitions.",
    "Is spooling supported, and how is it configured?",
    "What reports are predefined in this specification?",
    "What data variables are linked to the process completion event?",
    "Detail the parameter requirements for the remote command START.",
    "Are there any specific alarm IDs (ALIDs) linked to mass filter or beam drift?"
]


class EquipmentExtractor:
    _ENCODER_NAME = "cl100k_base"

    def __init__(self, llm_strategy: LLMStrategy) -> None:
        self._llm = llm_strategy.get_model(temperature=0, require_json=True)
        self._llm_retry = llm_strategy.get_model(temperature=0.2, require_json=True)
        self._encoder = tiktoken.get_encoding(self._ENCODER_NAME)
        self._chunk_tokens = settings.EXTRACTOR_CHUNK_TOKENS
        self._chunk_overlap_tokens = settings.EXTRACTOR_CHUNK_OVERLAP_TOKENS
        self._max_parallel = settings.EXTRACTOR_MAX_PARALLEL

    # ------------------------------------------------------------------
    # Public API
    # ------------------------------------------------------------------

    def extract_stage_1(
        self,
        pdf_text: str,
        section_csvs: dict[str, str],
        context_chunks: list[str] | None = None
    ) -> EquipmentSpec:
        """Stage 1: Extract Summary info and map standard SECS/GEM table columns via LLM, then build tables programmatically."""
        context_text = "\n\n".join(context_chunks) if context_chunks else pdf_text[:4000]
        
        table_dir_lines = []
        for section, csv_str in section_csvs.items():
            lines = csv_str.strip().split('\n')
            if lines:
                table_dir_lines.append(f"Table: {section}\nHeaders: {lines[0]}")
        
        table_directory = "\n\n".join(table_dir_lines)
        prompt = self._STAGE_1_PROMPT.format(context=context_text, table_directory=table_directory)
        
        response_obj = self._llm.invoke(prompt).content
        response_text = response_obj if isinstance(response_obj, str) else "".join(p.get("text", "") if isinstance(p, dict) else str(p) for p in response_obj)
        import json
        import csv
        import io
        from source.schemas.secsgem import SummarySpec, EquipmentSpec, StatusVariable, Event, Alarm, RemoteCommand, DataVariable, State, StateTransition
        from source.schemas.report import ReportDefinition

        try:
            # Clean possible markdown formatting
            cleaned = response_text.replace("```json", "").replace("```", "").strip()
            data = json.loads(cleaned)
        except Exception as e:
            logger.error("Failed to parse Stage 1 JSON: %s", e)
            data = {"Summary": {}, "ColumnMappings": {}}

        summary_data = data.get("Summary", {})
        summary = SummarySpec.model_validate(summary_data) if summary_data else None
        spec = EquipmentSpec(Summary=summary)
        if summary and summary.ToolID:
            spec.ToolID = summary.ToolID

        mappings = data.get("ColumnMappings", {})
        
        def build_entities(section_name, entity_class, list_attr):
            category_mappings = mappings.get(section_name)
            if not category_mappings:
                return
            if isinstance(category_mappings, dict):
                # Fallback for old schema if LLM hallucinated
                category_mappings = [{"TableKey": section_name, "Mapping": category_mappings}]
                
            entities = getattr(spec, list_attr, [])
            for item in category_mappings:
                table_key = item.get("TableKey")
                mapping = item.get("Mapping")
                if not table_key or not mapping or table_key not in section_csvs:
                    continue
                
                csv_str = section_csvs[table_key]
                reader = csv.DictReader(io.StringIO(csv_str))
                for row in reader:
                    obj_data = {}
                    for target_key, raw_col in mapping.items():
                        if raw_col in row:
                            obj_data[target_key] = row[raw_col]
                    if obj_data:
                        # Robust sanitization for Pydantic
                        for list_field in ["LinkedVIDs", "LinkedReports", "Parameters"]:
                            if list_field in obj_data and isinstance(obj_data[list_field], str):
                                val = obj_data[list_field].strip()
                                # handle multiline or comma separated
                                val = val.replace("\n", "").replace("\r", "")
                                obj_data[list_field] = [x.strip() for x in val.split(",") if x.strip()]
                        
                        if entity_class.__name__ == "State" and "Name" not in obj_data and "StateID" in obj_data:
                            obj_data["Name"] = str(obj_data["StateID"])
                        if entity_class.__name__ == "Alarm":
                            if "Name" not in obj_data and "AlarmID" in obj_data:
                                obj_data["Name"] = f"Alarm {obj_data['AlarmID']}"
                            # Fix non-integer AlarmIDs
                            if "AlarmID" in obj_data and isinstance(obj_data["AlarmID"], str) and not str(obj_data["AlarmID"]).isdigit():
                                if "Name" not in obj_data:
                                    obj_data["Name"] = obj_data["AlarmID"]
                                obj_data["AlarmID"] = abs(hash(obj_data["AlarmID"])) % 1000000
                        if entity_class.__name__ == "Event" and "Name" not in obj_data and "CEID" in obj_data:
                            obj_data["Name"] = f"Event {obj_data['CEID']}"
                        if entity_class.__name__ == "RemoteCommand":
                            if "Name" not in obj_data and "RCMD" in obj_data:
                                obj_data["Name"] = str(obj_data["RCMD"])
                            if "Parameters" in obj_data and isinstance(obj_data["Parameters"], list):
                                if len(obj_data["Parameters"]) > 0 and isinstance(obj_data["Parameters"][0], str):
                                    obj_data["Parameters"] = [{"Name": p, "Type": "UNKNOWN"} for p in obj_data["Parameters"]]
                            
                        try:
                            entities.append(entity_class.model_validate(obj_data))
                        except Exception as e:
                            logger.warning(f"Skipping row in {table_key}: {e}")
            setattr(spec, list_attr, entities)

        build_entities("StatusVariables", StatusVariable, "StatusVariables")
        build_entities("DataVariables", DataVariable, "DataVariables")
        build_entities("Events", Event, "Events")
        build_entities("Alarms", Alarm, "Alarms")
        build_entities("RemoteCommands", RemoteCommand, "RemoteCommands")
        build_entities("States", State, "States")
        build_entities("StateTransitions", StateTransition, "StateTransitions")
        build_entities("Reports", ReportDefinition, "Reports")
        
        return spec

    def extract_stage_2(
        self,
        spec: EquipmentSpec,
        pdf_text: str
    ) -> EquipmentSpec:
        """Stage 2: Chunk map-reduce deep Q&A."""
        chunks = self._chunk_text(pdf_text)
        workers = min(self._max_parallel, len(chunks))
        
        from concurrent.futures import ThreadPoolExecutor
        partial_specs: list[EquipmentSpec] = [None] * len(chunks)  # type: ignore[list-item]
        with ThreadPoolExecutor(max_workers=workers) as executor:
            futures = {
                executor.submit(self._extract_chunk, chunk, i + 1, len(chunks)): i
                for i, chunk in enumerate(chunks)
            }
            for future in futures:
                idx = futures[future]
                partial_specs[idx] = future.result()
                
        chunk_spec = self._merge_specs(partial_specs)
        merged = self._merge_specs([chunk_spec, spec])
        return merged
    
    # ------------------------------------------------------------------
    # Chunking (Split)
    # ------------------------------------------------------------------

    def _chunk_text(self, text: str) -> list[str]:
        """Split text into overlapping token-sized chunks."""
        tokens = self._encoder.encode(text)

        if len(tokens) <= self._chunk_tokens:
            return [text]

        chunks: list[str] = []
        start = 0
        stride = max(1, self._chunk_tokens - self._chunk_overlap_tokens)

        while start < len(tokens):
            end = start + self._chunk_tokens
            chunk_tokens = tokens[start:end]
            chunks.append(self._encoder.decode(chunk_tokens))
            if end >= len(tokens):
                break
            start += stride

        return chunks

    # ------------------------------------------------------------------
    # SML Script Extraction (deterministic, no LLM)
    # ------------------------------------------------------------------

    # Safety cap so a malformed/unterminated block can't run away to EOF.
    _SML_MAX_BLOCK_LINES = 400

    @staticmethod
    def _is_sml_terminator(stripped: str) -> bool:
        """True for a line that is solely the SML "." block terminator
        (optionally followed by a comment), matching TestScriptService."""
        return stripped == "." or (
            stripped.startswith(".")
            and (len(stripped) == 1 or stripped[1:].strip().startswith(("//", "#")))
        )

    def extract_sml_scripts(self, pdf_text: str) -> str:
        """Extract verbatim SML/SECS-II message blocks from the document text.

        This is a pure text parse — no LLM. SML samples in the manuals are
        self-delimiting: each begins with a bracketed header (e.g. "<S2F41 W")
        and ends either when its bracket nesting closes or at a lone "."
        terminator line. Each extracted block is normalised to end with a single
        "." so the output is directly consumable by
        ``TestScriptService.parse_sml_to_tests``. Returns the concatenated
        blocks, or an empty string when none are found.
        """
        if not pdf_text or not pdf_text.strip():
            return ""

        blocks = self._parse_sml_blocks(pdf_text)
        if not blocks:
            return ""

        logger.info("SML parse produced %d block(s).", len(blocks))
        return "\n\n".join(blocks)

    def _parse_sml_blocks(self, text: str) -> list[str]:
        """Deterministically pull every SML message block out of ``text``.

        A block starts at a bracketed ``<SxFy`` header line and runs until either
        bracket depth returns to zero (the message wrapper closed) or a lone "."
        terminator is reached — whichever comes first. Bare ``SxFy`` mentions in
        prose are ignored because they are not bracketed. Blocks are deduped on
        their normalised text, preserving document order.
        """
        lines = text.splitlines()
        n = len(lines)
        blocks: list[str] = []
        seen: set[str] = set()

        i = 0
        while i < n:
            if not _SML_HEADER_RE.match(lines[i].strip()):
                i += 1
                continue

            buf: list[str] = []
            depth = 0
            terminated = False
            j = i
            while j < n and (j - i) <= self._SML_MAX_BLOCK_LINES:
                line = lines[j]
                stripped = line.strip()
                buf.append(line)
                depth += line.count("<") - line.count(">")
                j += 1
                if self._is_sml_terminator(stripped):
                    terminated = True
                    break
                if depth <= 0:
                    terminated = True
                    break

            if terminated:
                # Drop trailing blank / "." lines, then normalise to one terminator.
                while buf and (not buf[-1].strip() or self._is_sml_terminator(buf[-1].strip())):
                    buf.pop()
                block = "\n".join(buf).rstrip()
                if block:
                    key = re.sub(r"\s+", " ", block).strip()
                    if key not in seen:
                        seen.add(key)
                        blocks.append(block + "\n.")
                # Skip any grouped terminator / blank lines following the block.
                i = j
                while i < n and (not lines[i].strip() or self._is_sml_terminator(lines[i].strip())):
                    i += 1
            else:
                i += 1

        return blocks

    # ------------------------------------------------------------------
    # Per-Chunk Extraction (Map)
    # ------------------------------------------------------------------

    def _extract_chunk(self, chunk: str, chunk_num: int, total_chunks: int) -> EquipmentSpec:
        """Extract from a single chunk. Returns an empty spec on failure in multi-chunk mode."""
        if total_chunks > 1:
            preamble = (
                f"This is section {chunk_num} of {total_chunks} from a larger document. "
                "Extract ONLY the SECS/GEM data present in THIS section. "
                "Return empty lists for any categories not found in this section.\n\n"
            )
        else:
            preamble = ""

        prompt = self._build_prompt(preamble + chunk)

        try:
            raw = self._llm.invoke(prompt).content
            if isinstance(raw, list):
                raw = "".join(part.get("text", "") if isinstance(part, dict) else str(part) for part in raw)
            data = json.loads(raw)
            if "$defs" in data:
                raise ValueError(
                    "You output the JSON schema itself instead of the extracted data."
                )
            self._sanitize(data)
            return EquipmentSpec.model_validate(data)
        except Exception as e:
            error_msg = str(e)
            logger.warning(
                "Primary extraction failed for chunk %d/%d (%s) - retrying.",
                chunk_num, total_chunks, error_msg,
            )
            retry_prompt = (
                prompt
                + f"\n\nCRITICAL ERROR IN PREVIOUS ATTEMPT:\n{error_msg}\n\n"
                "You MUST output a valid JSON object populated with the actual data "
                "from the document. DO NOT output the JSON schema definitions (e.g. '$defs')."
            )
            try:
                raw = self._llm_retry.invoke(retry_prompt).content
                if isinstance(raw, list):
                    raw = "".join(part.get("text", "") if isinstance(part, dict) else str(part) for part in raw)
                data = json.loads(raw)
                if "$defs" in data:
                    raise ValueError("Model persistently returned the schema instead of the data.")
                self._sanitize(data)
                return EquipmentSpec.model_validate(data)
            except Exception as retry_err:
                import traceback
                with open("extractor_error.txt", "w") as f:
                    f.write(traceback.format_exc())
                if total_chunks == 1:
                    raise
                logger.error(
                    "Chunk %d/%d failed after retry (%s). Returning empty spec for this chunk.",
                    chunk_num, total_chunks, retry_err,
                )
                return EquipmentSpec(ToolID="", ToolType="")

    # ------------------------------------------------------------------
    # Merging (Reduce)
    # ------------------------------------------------------------------

    @staticmethod
    def _merge_specs(specs: list[EquipmentSpec]) -> EquipmentSpec:
        merged = EquipmentSpec(ToolID="", ToolType="")

        for spec in specs:
            if not merged.DocumentType and spec.DocumentType:
                merged.DocumentType = spec.DocumentType
            if not merged.ToolID and spec.ToolID:
                merged.ToolID = spec.ToolID
            if not merged.ToolType and spec.ToolType:
                merged.ToolType = spec.ToolType
            if not merged.Model and spec.Model:
                merged.Model = spec.Model

        merged.StatusVariables = EquipmentExtractor._dedup_by_key(
            [sv for s in specs for sv in s.StatusVariables], key="SVID"
        )
        merged.DataVariables = EquipmentExtractor._dedup_by_key(
            [dv for s in specs for dv in s.DataVariables], key="DvID"
        )
        merged.Events = EquipmentExtractor._dedup_by_key(
            [ev for s in specs for ev in s.Events], key="CEID"
        )
        merged.Alarms = EquipmentExtractor._dedup_by_key(
            [al for s in specs for al in s.Alarms], key="AlarmID"
        )
        merged.RemoteCommands = EquipmentExtractor._dedup_by_key(
            [rc for s in specs for rc in s.RemoteCommands], key="RCMD"
        )
        merged.States = EquipmentExtractor._dedup_by_key(
            [st for s in specs for st in s.States], key="StateID"
        )
        merged.StateTransitions = EquipmentExtractor._dedup_transitions(
            [tr for s in specs for tr in s.StateTransitions]
        )
        
        # Merge Reports
        reports = []
        for s in specs:
            if hasattr(s, "Reports") and s.Reports:
                reports.extend(s.Reports)
        if reports:
            merged.Reports = EquipmentExtractor._dedup_by_key(reports, key="RPTID")
            
        # Merge Summary
        for spec in specs:
            if spec.Summary:
                if not merged.Summary:
                    merged.Summary = spec.Summary.model_copy(deep=True) if hasattr(spec.Summary, "model_copy") else spec.Summary.copy()
                else:
                    if not merged.Summary.EquipmentName and spec.Summary.EquipmentName:
                        merged.Summary.EquipmentName = spec.Summary.EquipmentName
                    if not merged.Summary.WaferSize and spec.Summary.WaferSize:
                        merged.Summary.WaferSize = spec.Summary.WaferSize
                    if not merged.Summary.SoftwareRevision and spec.Summary.SoftwareRevision:
                        merged.Summary.SoftwareRevision = spec.Summary.SoftwareRevision
                    if not merged.Summary.ToolID and spec.Summary.ToolID:
                        merged.Summary.ToolID = spec.Summary.ToolID
                    
                    # Merge arrays in summary
                    if spec.Summary.StandardsSupported:
                        if merged.Summary.StandardsSupported is None:
                            merged.Summary.StandardsSupported = []
                        merged.Summary.StandardsSupported.extend(spec.Summary.StandardsSupported)
                    if spec.Summary.GEMCompliance:
                        if merged.Summary.GEMCompliance is None:
                            merged.Summary.GEMCompliance = []
                        merged.Summary.GEMCompliance.extend(spec.Summary.GEMCompliance)
                    if not merged.Summary.HSMSConfiguration and spec.Summary.HSMSConfiguration:
                        merged.Summary.HSMSConfiguration = spec.Summary.HSMSConfiguration
                    if spec.Summary.StreamFunctions:
                        if merged.Summary.StreamFunctions is None:
                            merged.Summary.StreamFunctions = []
                        merged.Summary.StreamFunctions.extend(spec.Summary.StreamFunctions)
                    if spec.Summary.CommunicationStates:
                        if merged.Summary.CommunicationStates is None:
                            merged.Summary.CommunicationStates = []
                        merged.Summary.CommunicationStates.extend(spec.Summary.CommunicationStates)
                    if spec.Summary.ControlStates:
                        if merged.Summary.ControlStates is None:
                            merged.Summary.ControlStates = []
                        merged.Summary.ControlStates.extend(spec.Summary.ControlStates)

        return merged

    @staticmethod
    def _dedup_by_key(items: list, key: str) -> list:
        """Keep the highest-confidence entry per unique key value, merging list fields."""
        best: dict = {}
        for item in items:
            id_val = getattr(item, key)
            if id_val not in best:
                best[id_val] = item.model_copy(deep=True) if hasattr(item, "model_copy") else item.copy()
                continue

            existing = best[id_val]
            new_item = item
            
            conf_new = getattr(new_item, "Confidence", 1.0)
            conf_ext = getattr(existing, "Confidence", 1.0)

            if conf_new > conf_ext:
                existing, new_item = new_item.model_copy(deep=True) if hasattr(new_item, "model_copy") else new_item.copy(), existing

            if hasattr(existing, "model_fields"):
                for field in existing.model_fields.keys():
                    val_ext = getattr(existing, field)
                    val_new = getattr(new_item, field)

                    if isinstance(val_ext, list) and isinstance(val_new, list):
                        merged_list = []
                        for x in val_ext + val_new:
                            if x not in merged_list:
                                merged_list.append(x)
                        setattr(existing, field, merged_list)
                    elif isinstance(val_ext, str) and isinstance(val_new, str):
                        if val_ext in ("-", "", "unknown", "none") and val_new not in ("-", "", "unknown", "none"):
                            setattr(existing, field, val_new)
                        elif field == "Description" and len(val_new) > len(val_ext) and val_new not in ("-", "", "unknown", "none"):
                            setattr(existing, field, val_new)

            best[id_val] = existing

        return list(best.values())

    @staticmethod
    def _dedup_transitions(transitions: list) -> list:
        """Deduplicate state transitions by (FromState, ToState) tuple."""
        seen: dict[tuple[str, str], object] = {}
        for tr in transitions:
            key = (tr.FromState, tr.ToState)
            if key not in seen:
                seen[key] = tr
        return list(seen.values())

    # ------------------------------------------------------------------
    # Sanitization
    # ------------------------------------------------------------------

    @staticmethod
    def _sanitize(data: dict) -> None:
        """Clean up malformed transition entries before validation."""
        transitions = data.get("StateTransitions") or []
        data["StateTransitions"] = [
            t for t in transitions
            if t.get("FromState") and t.get("ToState")
        ]

    # ------------------------------------------------------------------
    # Prompt Builder
    # ------------------------------------------------------------------

    def _build_prompt(self, pdf_text: str) -> str:
        return f"""You extract semiconductor equipment SECS/GEM specs from technical documentation.

Output a single JSON object that conforms to this structure. No prose, no markdown fences, no commentary.
CRITICAL: DO NOT output the schema itself. You must output a JSON object populated with the actual extracted data.

EXPECTED JSON FORMAT:
{{
  "DocumentType": "string (User Manuals|Troubleshooting Guidance|GEM Manual|Variable Files|SML Scripts)",
  "ToolID": "string",
  "ToolType": "string",
  "Model": "string (optional)",
  "Protocol": "SECS/GEM",
  "Summary": {{
    "EquipmentName": "string - Extract from 'Project Name' if specified, else generic name",
    "WaferSize": "string",
    "SoftwareRevision": "string",
    "ToolID": "string - Extract from 'Project ID' if specified, else tool/machine ID",
    "StandardsSupported": [{{ "Standard": "string", "Version": "string" }}],
    "GEMCompliance": ["string"],
    "HSMSConfiguration": {{
      "DeviceID": "string",
      "IPAddress": "string",
      "PortNumber": "string",
      "BaudRate": "string",
      "Timeout": "string"
    }},
    "StreamFunctions": [{{ "Stream": 1, "Function": 1, "Description": "string" }}],
    "CommunicationStates": [{{ "State": "string", "Description": "string" }}],
    "ControlStates": [{{ "State": "string", "Description": "string" }}]
  }},
  "StatusVariables":[
    {{
      "SVID": 123,
      "Name": "string",
      "Description": "string",
      "DataType": "string",
      "AccessType": "string",
      "Value": "string",
      "Confidence": 0.0 to 1.0 (double)
    }}
  ],
  "DataVariables": [
    {{
      "DvID": 123,
      "Name": "string",
      "ValueType": "string (float|integer|string|boolean)",
      "Unit": "string",
      "Confidence": 0.0 to 1.0 (double)
    }}
  ],
  "Events": [
    {{
      "CEID": 123,
      "Name": "string",
      "Description": "string",
      "LinkedVIDs": [123],
      "LinkedReports": ["string"],
      "Confidence": 0.0 to 1.0 (double)
    }}
  ],
  "Alarms": [
    {{
      "AlarmID": 123,
      "Name": "string",
      "Severity": "critical|warning|info",
      "LinkedVID": 123,
      "Description": "string",
      "Confidence": 0.0 to 1.0
    }}
  ],
  "RemoteCommands": [
    {{
      "RCMD": "string - the name of the command/message",
      "Description": "string",
      "Parameters": [{{ "Name": "string", "Type": "string" }}],
      "Confidence": 0.0 to 1.0
    }}
  ],
  "States": [
    {{ "StateID": "string", "Name": "string", "Description": "string" }}
  ],
  "StateTransitions": [
    {{
      "FromState": "string",
      "ToState": "string",
      "TriggerEvent": "string",
      "TriggerCommand": "string",
      "Manual": false
    }}
  ]
}}

CONFIDENCE RUBRIC (set the `Confidence` field per-entity):
- 1.0  = verbatim from a clearly labeled table
- 0.7  = inferred from prose with a clear section heading or strong signal
- <0.5 = guessed from weak context

HARD RULES:
- If the document explicitly provides "Project ID" and "Project Name", map them to `Summary.ToolID` and `Summary.EquipmentName` respectively.
- For every Event, you MUST extract the LinkedVIDs (Variables) associated with that event.
- Populate `LinkedVIDs` from any "Linked VIDs" column or comma-separated VID list following an event description.
- Populate `StateTransitions` from text matching `<State> -> <State> (Triggered by <event/command>)` patterns.
  - `FromState` and `ToState` are REQUIRED, must be non-null state names matching entries in `States`.
  - If the trigger is an event, set `TriggerEvent` (the event name) and leave `TriggerCommand` null.
  - If the trigger is a command, set `TriggerCommand` (the RCMD name) and leave `TriggerEvent` null.
  - If the transition is described as "manual" with no event/command, set `Manual: true` and leave both trigger fields null.
  - Never set both `TriggerEvent` and `TriggerCommand`.
- `Severity` MUST be lowercase: `critical`, `warning`, or `info`.
- CRITICAL: DO NOT hallucinate standard GEM capabilities (e.g., "Establish Communications", "Spooling", "Alarm Management") as Events or Alarms. ONLY extract entities that are explicitly defined with specific IDs in the document.
- If a section is absent from the document, return an empty list. Do NOT invent IDs, names, or fields.
- Output a single JSON object that validates against the schema above. Nothing else.

### DOCUMENT
{pdf_text}
"""

    # ------------------------------------------------------------------
    # Table Extraction (PDF → raw CSV → LLM)
    # ------------------------------------------------------------------
    _COMBINED_TABLE_PROMPT =   """You are a SECS/GEM integration expert.
The tables below were extracted from a semiconductor equipment manual.
Each table is labelled with its SECS/GEM section type.
Extract every row from every table and return a SINGLE JSON object.

RULES:
- StatusVariables  : {{ "SVID": int, "Name": str, "Description": str, "DataType": str, "AccessType": str, "Value": str, "Confidence": float }}
- DataVariables    : {{ "DvID": int, "Name": str, "ValueType": "float|integer|string|boolean", "Unit": str }}
- Events           : {{ "CEID": int, "Name": str, "Description": str, "LinkedVIDs": [int], "LinkedReports": ["str"], "Confidence": float }}
                     SKIP any row without a numeric CEID. Parse comma-separated VID lists into LinkedVIDs.
- Alarms           : {{ "AlarmID": int, "Name": str, "Severity": "critical|warning|info", "LinkedVID": int|null, "Description": str, "Confidence": float }}
                     SKIP any row without a numeric AlarmID. Severity MUST be lowercase.
- RemoteCommands   : {{ "RCMD": str, "Description": str, "Parameters": [{{"Name": str, "Type": str}}], "Confidence": float }}
- States           : {{ "StateID": str, "Name": str, "Description": str }}
- StateTransitions : {{ "FromState": str (REQUIRED), "ToState": str (REQUIRED), "TriggerEvent": str|null, "TriggerCommand": str|null, "Manual": bool }}
                     Never set both TriggerEvent and TriggerCommand on the same entry.
- Reports          : {{ "RPTID": str, "Name": str, "LinkedVIDs": [int] }}

Return ONLY this JSON structure (include only sections that have tables):
{{
  "StatusVariables": [...],
  "DataVariables": [...],
  "Events": [...],
  "Alarms": [...],
  "RemoteCommands": [...],
  "States": [...],
  "StateTransitions": [...],
  "Reports": [...]
}}

No prose, no markdown fences.

{tables}
"""

    # Char budget for the combined table call.
    # 56k chars ≈ 14k tokens — conservative enough to fit any typical GEM ICD.
    # Raise this if your LLM context allows and your tables are large.
    _COMBINED_TABLE_CHAR_LIMIT = 56_000

    _SHEET_NAME_KEYWORDS: dict[str, set[str]] = {
        "StatusVariables": {"svid", "sv", "status variable", "status_var", "status"},
        "DataVariables": {"dvid", "dv", "data variable", "data_var"},
        "Events": {"ceid", "ce", "collection event", "event"},
        "Alarms": {"alarm", "alid"},
        "RemoteCommands": {"rcmd", "remote command", "command"},
        "States": {"state"},
        "StateTransitions": {"transition", "state machine"},
        "Reports": {"report", "rptid", "reports"},
    }

    _SECTION_KEYWORDS: dict[str, set[str]] = {
        "StatusVariables": {
            "svid", "sv id", "status variable", "statusvariable",
            "variable id", "variable name", "access type",
        },
        "DataVariables": {
            "dvid", "dv id", "data variable", "datavariable",
            "value type", "valuetype",
        },
        "Events": {
            "ceid", "ce id", "collection event", "collectionevent",
            "event id", "event name", "linked vid",
        },
        "Alarms": {
            "alarm id", "alarmid", "alarm name", "alarm text",
            "alarm code", "alid", "al id", "severity",
        },
        "RemoteCommands": {
            "rcmd", "remote command", "remotecommand", "command name",
            "command id", "command",
        },
        "States": {
            "state id", "stateid", "state name", "equipment state",
        },
        "StateTransitions": {
            "from state", "to state", "fromstate", "tostate",
            "transition", "trigger",
        },
        "Reports": {
            "report id", "rptid", "rpt id", "report name",
        },
    }

    _CSV_FILENAMES: dict[str, str] = {
        "StatusVariables": "status_variables.csv",
        "DataVariables": "data_variables.csv",
        "Events": "events.csv",
        "Alarms": "alarms.csv",
        "RemoteCommands": "remote_commands.csv",
        "States": "states.csv",
        "StateTransitions": "state_transitions.csv",
        "Reports": "reports.csv",
    }

    _STAGE_1_PROMPT: str = """You are an expert semiconductor equipment integration engineer.
We have extracted tables from an equipment manual.

CONTEXT OVERVIEW (for basic info like ToolID, Name, Software Version):
{context}

TABLE DIRECTORY:
The following tables were extracted. For each table, we provide its Section Name and its Column Headers.
{table_directory}

YOUR TASK:
1. Extract the "Summary" information (Equipment Name, Standards, HSMS Config, etc.) using the Context Overview and by identifying which dynamic tables contain the required information.
2. For the standard SECS/GEM tables, map the raw column headers to our standard schema keys.
Our Standard Schema Keys:
- StatusVariables: SVID, Name, Description, DataType, AccessType, Value
- DataVariables: DvID, Name, ValueType, Unit
- Events: CEID, Name, Description, LinkedVIDs, LinkedReports
- Alarms: AlarmID, Name, Severity, LinkedVID, Description
- RemoteCommands: RCMD, Description, Parameters
- States: StateID, Name, Description
- StateTransitions: FromState, ToState, TriggerEvent, TriggerCommand
- Reports: RPTID, Name, LinkedVIDs

Return ONLY valid JSON matching the following structure:
{{
    "Summary": {{
        "EquipmentName": "string",
        "WaferSize": "string",
        "SoftwareRevision": "string",
        "ToolID": "string",
        "StandardsSupported": [{{"Standard": "string", "Version": "string"}}],
        "GEMCompliance": ["string"],
        "HSMSConfiguration": {{"DeviceID": "string", "IPAddress": "string", "PortNumber": "string", "BaudRate": "string", "Timeout": "string"}},
        "StreamFunctions": [{{"Stream": "string", "Function": "string", "Description": "string"}}],
        "CommunicationStates": [{{"State": "string", "Description": "string"}}],
        "ControlStates": [{{"State": "string", "Description": "string"}}]
    }},
    "ColumnMappings": {{
        "StatusVariables": [{{"TableKey": "exact name from TABLE DIRECTORY", "Mapping": {{"SVID": "raw column name", ...}}}}],
        "Events": [{{"TableKey": "exact name from TABLE DIRECTORY", "Mapping": {{"CEID": "raw column name", ...}}}}],
        "Alarms": [{{"TableKey": "exact name from TABLE DIRECTORY", "Mapping": {{"AlarmID": "raw column name", ...}}}}],
        "DataVariables": [{{"TableKey": "exact name from TABLE DIRECTORY", "Mapping": {{"DvID": "raw column name", ...}}}}],
        "RemoteCommands": [{{"TableKey": "exact name from TABLE DIRECTORY", "Mapping": {{"RCMD": "raw column name", ...}}}}],
        "States": [{{"TableKey": "exact name from TABLE DIRECTORY", "Mapping": {{"StateID": "raw column name", ...}}}}],
        "StateTransitions": [{{"TableKey": "exact name from TABLE DIRECTORY", "Mapping": {{"FromState": "raw column name", ...}}}}],
        "Reports": [{{"TableKey": "exact name from TABLE DIRECTORY", "Mapping": {{"RPTID": "raw column name", ...}}}}]
    }}
}}
No prose, no markdown fences.
"""

    _TABLE_PROMPTS: dict[str, str] = {
        "StatusVariables": """You are a SECS/GEM integration expert.
The table below was extracted directly from a semiconductor equipment manual.
The column names may differ from standard SECS/GEM field names — map them intelligently.
Convert every data row into a JSON array of StatusVariable objects.
Each object MUST have:
  "SVID": integer, "Name": string, "Description": string (empty if absent),
  "DataType": string (e.g. "U4","ASCII","Float","String"),
  "AccessType": string (e.g. "RO","RW"), "Value": string (empty if absent),
  "Confidence": float (1.0 if SVID+Name present else 0.7)
Return ONLY: {{"StatusVariables": [...]}}  No prose, no markdown fences.
TABLE (CSV):
{csv}""",

        "DataVariables": """You are a SECS/GEM integration expert.
The table below was extracted directly from a semiconductor equipment manual.
The column names may differ from standard SECS/GEM field names — map them intelligently.
Convert every data row into a JSON array of DataVariable objects.
Each object MUST have:
  "DvID": integer, "Name": string,
  "ValueType": string (float|integer|string|boolean),
  "Unit": string (empty if absent)
Return ONLY: {{"DataVariables": [...]}}  No prose, no markdown fences.
TABLE (CSV):
{csv}""",

        "Events": """You are a SECS/GEM integration expert.
The table below was extracted directly from a semiconductor equipment manual.
The column names may differ from standard SECS/GEM field names — map them intelligently.
Convert every data row into a JSON array of Collection Event objects.
CRITICAL: If a row does not contain an explicit numeric ID (CEID), IGNORE IT completely. Do NOT invent IDs for standard capabilities.
Each object MUST have:
  "CEID": integer, "Name": string, "Description": string (empty if absent),
  "LinkedVIDs": [integer] (parse comma-separated VID lists, empty list if absent),
  "LinkedReports": ["string"] (parse explicit report IDs if mentioned, else empty),
  "Confidence": float (1.0 if CEID+Name present else 0.7)
Return ONLY: {{"Events": [...]}}  No prose, no markdown fences.
TABLE (CSV):
{csv}""",

        "Alarms": """You are a SECS/GEM integration expert.
The table below was extracted directly from a semiconductor equipment manual.
The column names may differ from standard SECS/GEM field names — map them intelligently.
Convert every data row into a JSON array of Alarm objects.
CRITICAL: If a row does not contain an explicit numeric ID (AlarmID), IGNORE IT completely. Do NOT invent IDs.
Each object MUST have:
  "AlarmID": integer, "Name": string,
  "Severity": exactly one of "critical"|"warning"|"info" (lowercase),
  "LinkedVID": integer or null, "Description": string (empty if absent),
  "Confidence": float (1.0 if AlarmID+Name present else 0.7)
Return ONLY: {{"Alarms": [...]}}  No prose, no markdown fences.
TABLE (CSV):
{csv}""",

        "RemoteCommands": """You are a SECS/GEM integration expert.
The table below was extracted directly from a semiconductor equipment manual.
The column names may differ from standard SECS/GEM field names — map them intelligently.
Convert every data row into a JSON array of RemoteCommand objects.
Each object MUST have:
  "RCMD": string, "Description": string (empty if absent),
  "Parameters": [{{"Name": string, "Type": string}}] (parse from Parameters column, empty list if None/absent),
  "Confidence": float (1.0 if command name present else 0.7)
Return ONLY: {{"RemoteCommands": [...]}}  No prose, no markdown fences.
TABLE (CSV):
{csv}""",

        "States": """You are a SECS/GEM integration expert.
The table below was extracted directly from a semiconductor equipment manual.
Convert every data row into a JSON array of State objects.
Each object MUST have:
  "StateID": string, "Name": string, "Description": string (empty if absent)
Return ONLY: {{"States": [...]}}  No prose, no markdown fences.
TABLE (CSV):
{csv}""",

        "StateTransitions": """You are a SECS/GEM integration expert.
The table below was extracted directly from a semiconductor equipment manual.
The column names may differ — map them intelligently.
Convert every data row into a JSON array of StateTransition objects.
Each object MUST have:
  "FromState": string (REQUIRED), "ToState": string (REQUIRED),
  "TriggerEvent": string or null, "TriggerCommand": string or null,
  "Manual": boolean (true only if no event/command trigger)
Never set both TriggerEvent and TriggerCommand on the same entry.
Return ONLY: {{"StateTransitions": [...]}}  No prose, no markdown fences.
TABLE (CSV):
{csv}""",

        "Reports": """You are a SECS/GEM integration expert.
The table below was extracted directly from a semiconductor equipment manual.
The column names may differ from standard SECS/GEM field names — map them intelligently.
Convert every data row into a JSON array of ReportDefinition objects.
Each object MUST have:
  "RPTID": string (or integer), "Name": string,
  "LinkedVIDs": [integer] (parse comma-separated VID lists, empty list if absent)
Return ONLY: {{"Reports": [...]}}  No prose, no markdown fences.
TABLE (CSV):
{csv}""",
    }
    def _extract_all_tables_combined(
        self, section_csvs: dict[str, str]
    ) -> "EquipmentSpec | None":
        """Try to extract ALL classified table sections in one LLM call.

        Returns a partial EquipmentSpec on success, or None when the caller
        should fall back to per-section calls.  Fails gracefully on token
        overflow, JSON parse error, or any exception.
        """
        if not section_csvs:
            return None

        # Only send standard SECS/GEM tables to the LLM
        standard_sections = {
            "StatusVariables", "DataVariables", "Events", "Alarms", 
            "RemoteCommands", "States", "StateTransitions", "Reports"
        }
        
        table_blocks = []
        for section, csv_str in section_csvs.items():
            if section in standard_sections:
                table_blocks.append(f"--- TABLE: {section} ---\n{csv_str}")
        
        if not table_blocks:
            return None
            
        tables_text = "\n\n".join(table_blocks)

        prompt = self._COMBINED_TABLE_PROMPT.format(tables=tables_text)

        if len(prompt) > self._COMBINED_TABLE_CHAR_LIMIT:
            logger.info(
                "Combined table prompt is %d chars (limit %d) — falling back to per-table.",
                len(prompt), self._COMBINED_TABLE_CHAR_LIMIT,
            )
            return None

        for attempt, llm in enumerate((self._llm, self._llm_retry), start=1):
            try:
                raw = llm.invoke(prompt).content
                if isinstance(raw, list):
                    raw = "".join(
                        p.get("text", "") if isinstance(p, dict) else str(p) for p in raw
                    )
                return self._parse_combined_table_response(raw)
            except Exception as exc:
                logger.warning("Combined table attempt %d failed: %s", attempt, exc)

        logger.warning("Combined table extraction failed after retry — using per-table fallback.")
        return None

    def _parse_combined_table_response(self, raw: str) -> "EquipmentSpec | None":
        """Parse the LLM JSON from a combined table call into a partial EquipmentSpec."""
        cleaned = raw.strip()
        if cleaned.startswith("```"):
            lines = cleaned.splitlines()
            cleaned = "\n".join(
                line for line in lines if not line.strip().startswith("```")
            ).strip()

        try:
            data = json.loads(cleaned)
        except json.JSONDecodeError as exc:
            logger.warning("Combined table response is not valid JSON: %s", exc)
            return None

        spec = EquipmentSpec(ToolID="", ToolType="")
        section_map = {
            "StatusVariables":  ("StatusVariables",  StatusVariable),
            "DataVariables":    ("DataVariables",     DataVariable),
            "Events":           ("Events",            Event),
            "Alarms":           ("Alarms",            Alarm),
            "RemoteCommands":   ("RemoteCommands",    RemoteCommand),
            "States":           ("States",            State),
            "StateTransitions": ("StateTransitions",  StateTransition),
        }
        for key, (attr, model) in section_map.items():
            items = data.get(key) or []
            if not items:
                continue
            parsed = []
            for item in items:
                try:
                    if key == "StateTransitions" and not (item.get("FromState") and item.get("ToState")):
                        continue
                    parsed.append(model.model_validate(item))
                except Exception as exc:
                    logger.warning("Skipping invalid %s item %s: %s", key, item, exc)
            setattr(spec, attr, parsed)

        report_items = data.get("Reports") or []
        if report_items:
            from source.schemas.report import ReportDefinition
            reports = []
            for item in report_items:
                try:
                    reports.append(ReportDefinition.model_validate(item))
                except Exception as exc:
                    logger.warning("Skipping invalid Report item: %s", exc)
            spec.Reports = reports

        logger.info(
            "Combined table parse: %d SVs, %d DVs, %d Events, %d Alarms, %d RCMDs",
            len(spec.StatusVariables), len(spec.DataVariables),
            len(spec.Events), len(spec.Alarms), len(spec.RemoteCommands),
        )
        return spec

    def _classify_table(self, rows: list[list[str]]) -> str | None:
        """Return the section name that best matches the table header, or None."""
        if not rows:
            return None
        header_text = " ".join(rows[0]).lower()
        best_section: str | None = None
        best_hits = 0
        for section, keywords in self._SECTION_KEYWORDS.items():
            hits = sum(1 for kw in keywords if kw in header_text)
            if hits > best_hits:
                best_hits = hits
                best_section = section
        return best_section if best_hits > 0 else None

    def _classify_by_sheet_name(self, sheet_name: str) -> str | None:
        """Return the section name that best matches an Excel sheet name, or None."""
        name_lower = sheet_name.lower()
        for section, keywords in self._SHEET_NAME_KEYWORDS.items():
            if any(kw in name_lower for kw in keywords):
                return section
        return None

    # ------------------------------------------------------------------
    # Excel Extraction
    # ------------------------------------------------------------------

    def extract_excel(self, excel_path: Union[str, Path]) -> EquipmentSpec:
        """Extract a SECS/GEM EquipmentSpec from an Excel workbook.

        Each sheet is converted to CSV and fed to the same per-section LLM
        prompts used for PDF table extraction. Sheets are classified first by
        their column headers, then by sheet name as fallback.
        """
        import openpyxl

        excel_path = Path(excel_path)
        try:
            wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
        except Exception as exc:
            logger.error("Failed to open Excel file %s: %s", excel_path.name, exc)
            return EquipmentSpec(ToolID="", ToolType="")

        partial_specs: list[EquipmentSpec] = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = [
                [str(cell.value).strip() if cell.value is not None else "" for cell in row]
                for row in ws.iter_rows()
            ]
            rows = [row for row in rows if any(cell for cell in row)]
            if len(rows) < 2:
                continue

            section = self._classify_table(rows) or self._classify_by_sheet_name(sheet_name)
            if not section:
                logger.info("Sheet '%s' could not be classified — skipping.", sheet_name)
                continue

            buf = io.StringIO()
            writer = csv.writer(buf)
            writer.writerows(rows)
            csv_str = buf.getvalue()

            logger.info("Extracting sheet '%s' as section '%s'", sheet_name, section)
            partial = self._extract_from_csv(section, csv_str)
            if partial:
                partial_specs.append(partial)

        wb.close()

        if not partial_specs:
            logger.warning("No classifiable sheets found in %s", excel_path.name)
            return EquipmentSpec(ToolID="", ToolType="")

        merged = self._merge_specs(partial_specs)
        logger.info(
            "Excel extraction done: %d SVs, %d DVs, %d Events, %d Alarms",
            len(merged.StatusVariables), len(merged.DataVariables),
            len(merged.Events), len(merged.Alarms),
        )
        return merged

    def extract_and_save_tables(
        self,
        pdf_path: Path,
        tables_dir: Union[Path, None],
       
        tables_store_path: Union[Path, None] = None,
    ) -> dict[str, str]:
        """
        Extract all tables from the PDF using pdfplumber.
        Classify each table by section.
        Save each classified table as-is to a CSV file (raw column names preserved).
        Return a dict of {section: csv_string} for LLM consumption.
        """
        section_rows: dict[str, list[list[str]]] = {}

        try:
            with pdfplumber.open(pdf_path) as pdf:
                table_index = 0
                for page_num, page in enumerate(pdf.pages, 1):
                    for table in page.find_tables():
                        table_index += 1
                        raw_table = table.extract()
                        cleaned = [
                            [str(cell).strip() if cell is not None else "" for cell in row]
                            for row in raw_table
                            if any(cell not in [None, "", " "] for cell in row)
                        ]
                        if len(cleaned) < 2:
                            continue
                        
                        section = self._classify_table(cleaned)
                        if not section:
                            # Attempt to find heading above table
                            x0, top, x1, bottom = table.bbox
                            crop_top = max(0, top - 100) # Look back up to 100 points
                            heading = ""
                            if crop_top < top:
                                heading_bbox = (0, crop_top, page.width, top)
                                heading_crop = page.within_bbox(heading_bbox)
                                text = heading_crop.extract_text()
                                if text:
                                    # Get last non-empty line
                                    lines = [line.strip() for line in text.split('\n') if line.strip()]
                                    if lines:
                                        heading = lines[-1]
                            
                            if heading:
                                # Sanitize heading
                                import re
                                section = re.sub(r'[^A-Za-z0-9_]', '_', heading)
                                section = re.sub(r'_+', '_', section).strip('_')
                                
                            if not section:
                                section = f"Unclassified_Page{page_num}_Table{table_index}"

                        if section not in section_rows:
                            section_rows[section] = cleaned
                        else:
                            # Compare headers to ensure we don't mix different tables
                            existing_header = section_rows[section][0]
                            new_header = cleaned[0]
                            if len(existing_header) == len(new_header) and existing_header == new_header:
                                # Safe to append (likely a multi-page table)
                                section_rows[section].extend(cleaned[1:])
                            else:
                                # Headers don't match! It's a completely different table.
                                idx = 2
                                while f"{section}_{idx}" in section_rows:
                                    idx += 1
                                section_rows[f"{section}_{idx}"] = cleaned
        except Exception as exc:
            logger.error("pdfplumber failed on %s: %s", pdf_path.name, exc)
            return {}

        if not section_rows:
            logger.info("No classifiable tables found in %s", pdf_path.name)
            return {}

        section_csvs: dict[str, str] = {}

        for section, rows in section_rows.items():
            # Serialise to CSV string
            buf = io.StringIO()
            writer = csv.writer(buf)
            writer.writerows(rows)
            csv_str = buf.getvalue()
            section_csvs[section] = csv_str

            # Save to disk if tables_dir provided
            if tables_dir is not None:
                raw_tables_dir = tables_dir / "RawTables"
                raw_tables_dir.mkdir(parents=True, exist_ok=True)
                filename = self._CSV_FILENAMES.get(section, f"{section}.csv")
                csv_path = raw_tables_dir / filename
                # Append new data rows to existing file, or create fresh
                if csv_path.exists():
                    existing_lines = csv_path.read_text(encoding="utf-8").splitlines()
                    new_lines = csv_str.splitlines()
                    # Skip header of incoming if file already has content
                    combined = existing_lines + new_lines[1:] if existing_lines else new_lines
                    csv_path.write_text("\n".join(combined) + "\n", encoding="utf-8")
                else:
                    csv_path.write_text(csv_str, encoding="utf-8")
                logger.info("Saved raw %s table (%d rows) to %s", section, len(rows), csv_path)

        # ── Index table rows into the dedicated 'tables' vector store ──────────
        if tables_store_path is not None and section_rows:
            try:
                from source.utils.embedder import VectorStoreManager
                tables_vs = VectorStoreManager(tables_store_path)
                for section, rows in section_rows.items():
                    if len(rows) < 2:
                        continue
                    headers = rows[0]
                    for row_idx, row in enumerate(rows[1:], start=1):
                        parts = []
                        for header, cell in zip(headers, row):
                            if cell:
                                parts.append(f"{header}: {cell}")
                        if not parts:
                            continue
                        sentence = f"[{section}] " + " | ".join(parts)
                        tables_vs.add_document(
                            sentence,
                            metadata={
                                "project_id": str(pdf_path.parent.parent.name),
                                "document_id": pdf_path.stem,
                                "document_category": "tables",
                                "section": section,
                                "row_index": row_idx,
                            },
                        )
                logger.info(
                    "Indexed table rows from %s into tables vector store at %s",
                    pdf_path.name, tables_store_path,
                )
            except Exception as exc:
                logger.warning(
                    "Table vector store indexing failed for %s (non-fatal): %s",
                    pdf_path.name, exc,
                )

        return section_csvs

    def _extract_from_csv(self, section: str, csv_str: str) -> EquipmentSpec | None:
        """Send a raw CSV table to the LLM with a section-specific prompt."""
        prompt_template = self._TABLE_PROMPTS.get(section)
        if not prompt_template:
            return None

        prompt = prompt_template.format(csv=csv_str)

        for attempt, llm in enumerate((self._llm, self._llm_retry), start=1):
            try:
                raw = llm.invoke(prompt).content
                if isinstance(raw, list):
                    raw = "".join(
                        p.get("text", "") if isinstance(p, dict) else str(p) for p in raw
                    )
                return self._parse_table_response(raw, section)
            except Exception as exc:
                logger.warning(
                    "Table LLM attempt %d failed for section %s: %s", attempt, section, exc
                )

        logger.error("Both LLM attempts failed for section %s — skipping.", section)
        return None

    def _parse_table_response(self, raw: str, section: str) -> EquipmentSpec | None:
        """Parse LLM JSON response for a single table section into a partial EquipmentSpec."""
        try:
            data = json.loads(raw)
        except json.JSONDecodeError as exc:
            logger.warning("JSON parse failed for %s table response: %s", section, exc)
            return None

        spec = EquipmentSpec(ToolID="", ToolType="")
        try:
            items = data.get(section) or []
            if section == "StatusVariables":
                spec.StatusVariables = [StatusVariable.model_validate(i) for i in items]
            elif section == "DataVariables":
                spec.DataVariables = [DataVariable.model_validate(i) for i in items]
            elif section == "Events":
                spec.Events = [Event.model_validate(i) for i in items]
            elif section == "Alarms":
                spec.Alarms = [Alarm.model_validate(i) for i in items]
            elif section == "RemoteCommands":
                spec.RemoteCommands = [RemoteCommand.model_validate(i) for i in items]
            elif section == "States":
                spec.States = [State.model_validate(i) for i in items]
            elif section == "StateTransitions":
                valid = [t for t in items if t.get("FromState") and t.get("ToState")]
                spec.StateTransitions = [StateTransition.model_validate(t) for t in valid]
            elif section == "Reports":
                from source.schemas.secsgem import ReportDefinition
                spec.Reports = [ReportDefinition.model_validate(i) for i in items]
        except Exception as exc:
            logger.warning("model_validate failed for %s table response: %s", section, exc)
            return None

        return spec
