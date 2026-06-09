import csv
import io
import json
import logging
from concurrent.futures import ThreadPoolExecutor
from pathlib import Path
from typing import Union

import pdfplumber
import tiktoken

from source.config import settings
from source.schemas.secsgem import (
    Alarm,
    DataVariable,
    EquipmentSpec,
    Event,
    RemoteCommand,
    RCMDParameter,
    State,
    StateTransition,
    StatusVariable,
)
from source.utils.llm_factory import LLMStrategy

logger = logging.getLogger(__name__)


class EquipmentExtractor:
    _ENCODER_NAME = "cl100k_base"

    def __init__(self, llm_strategy: LLMStrategy) -> None:
        self._llm = llm_strategy.get_model(temperature=0, require_json=True)
        self._llm_retry = llm_strategy.get_model(temperature=0.2, require_json=True)
        self._encoder = tiktoken.get_encoding(self._ENCODER_NAME)
        self._chunk_tokens = settings.EXTRACTOR_CHUNK_TOKENS
        self._chunk_overlap_tokens = settings.EXTRACTOR_CHUNK_OVERLAP_TOKENS
        self._max_parallel = settings.EXTRACTOR_MAX_PARALLEL

    # ------------------------------------------------------------------
    # Public API
    # ------------------------------------------------------------------

    def extract(self, pdf_text: str, section_csvs: dict[str, str] = None) -> EquipmentSpec:
        section_csvs = section_csvs or {}

        # Step 2: Text chunks (always run — gives LLM overall document context)
        chunks = self._chunk_text(pdf_text)

        if len(chunks) == 1:
            chunk_spec = self._extract_chunk(chunks[0], chunk_num=1, total_chunks=1)
        else:
            workers = min(self._max_parallel, len(chunks))
            logger.info(
                "Document split into %d chunks; extracting up to %d in parallel.",
                len(chunks), workers,
            )
            partial_specs: list[EquipmentSpec] = [None] * len(chunks)  # type: ignore[list-item]
            with ThreadPoolExecutor(max_workers=workers) as executor:
                futures = {
                    executor.submit(self._extract_chunk, chunk, i + 1, len(chunks)): i
                    for i, chunk in enumerate(chunks)
                }
                for future in futures:
                    idx = futures[future]
                    partial_specs[idx] = future.result()
            chunk_spec = self._merge_specs(partial_specs)

        # Step 3: If we have classified table CSVs, extract each section precisely
        # and merge on top of the chunk-based spec (table results win on conflict)
        if section_csvs:
            table_partials: list[EquipmentSpec] = [chunk_spec]
            for section, csv_str in section_csvs.items():
                partial = self._extract_from_csv(section, csv_str)
                if partial:
                    table_partials.append(partial)
            merged = self._merge_specs(table_partials)
        else:
            merged = chunk_spec

        logger.info(
            "Final spec: %d SVs, %d DVs, %d Events, %d Alarms, %d RCMDs",
            len(merged.StatusVariables),
            len(merged.DataVariables),
            len(merged.Events),
            len(merged.Alarms),
            len(merged.RemoteCommands),
        )
        return merged

    # ------------------------------------------------------------------
    # Chunking (Split)
    # ------------------------------------------------------------------

    def _chunk_text(self, text: str) -> list[str]:
        """Split text into overlapping token-sized chunks."""
        tokens = self._encoder.encode(text)

        if len(tokens) <= self._chunk_tokens:
            return [text]

        chunks: list[str] = []
        start = 0
        stride = self._chunk_tokens - self._chunk_overlap_tokens

        while start < len(tokens):
            end = start + self._chunk_tokens
            chunk_tokens = tokens[start:end]
            chunks.append(self._encoder.decode(chunk_tokens))
            if end >= len(tokens):
                break
            start += stride

        return chunks

    # ------------------------------------------------------------------
    # Per-Chunk Extraction (Map)
    # ------------------------------------------------------------------

    def _extract_chunk(self, chunk: str, chunk_num: int, total_chunks: int) -> EquipmentSpec:
        """Extract from a single chunk. Returns an empty spec on failure in multi-chunk mode."""
        if total_chunks > 1:
            preamble = (
                f"This is section {chunk_num} of {total_chunks} from a larger document. "
                "Extract ONLY the SECS/GEM data present in THIS section. "
                "Return empty lists for any categories not found in this section.\n\n"
            )
        else:
            preamble = ""

        prompt = self._build_prompt(preamble + chunk)

        try:
            raw = self._llm.invoke(prompt).content
            if isinstance(raw, list):
                raw = "".join(part.get("text", "") if isinstance(part, dict) else str(part) for part in raw)
            data = json.loads(raw)
            if "$defs" in data:
                raise ValueError(
                    "You output the JSON schema itself instead of the extracted data."
                )
            self._sanitize(data)
            return EquipmentSpec.model_validate(data)
        except Exception as e:
            error_msg = str(e)
            logger.warning(
                "Primary extraction failed for chunk %d/%d (%s) - retrying.",
                chunk_num, total_chunks, error_msg,
            )
            retry_prompt = (
                prompt
                + f"\n\nCRITICAL ERROR IN PREVIOUS ATTEMPT:\n{error_msg}\n\n"
                "You MUST output a valid JSON object populated with the actual data "
                "from the document. DO NOT output the JSON schema definitions (e.g. '$defs')."
            )
            try:
                raw = self._llm_retry.invoke(retry_prompt).content
                if isinstance(raw, list):
                    raw = "".join(part.get("text", "") if isinstance(part, dict) else str(part) for part in raw)
                data = json.loads(raw)
                if "$defs" in data:
                    raise ValueError("Model persistently returned the schema instead of the data.")
                self._sanitize(data)
                return EquipmentSpec.model_validate(data)
            except Exception as retry_err:
                if total_chunks == 1:
                    raise
                logger.error(
                    "Chunk %d/%d failed after retry (%s). Returning empty spec for this chunk.",
                    chunk_num, total_chunks, retry_err,
                )
                return EquipmentSpec(ToolID="", ToolType="")

    # ------------------------------------------------------------------
    # Merging (Reduce)
    # ------------------------------------------------------------------

    @staticmethod
    def _merge_specs(specs: list[EquipmentSpec]) -> EquipmentSpec:
        merged = EquipmentSpec(ToolID="", ToolType="")

        for spec in specs:
            if not merged.DocumentType and spec.DocumentType:
                merged.DocumentType = spec.DocumentType
            if not merged.ToolID and spec.ToolID:
                merged.ToolID = spec.ToolID
            if not merged.ToolType and spec.ToolType:
                merged.ToolType = spec.ToolType
            if not merged.Model and spec.Model:
                merged.Model = spec.Model

        merged.StatusVariables = EquipmentExtractor._dedup_by_key(
            [sv for s in specs for sv in s.StatusVariables], key="SVID"
        )
        merged.DataVariables = EquipmentExtractor._dedup_by_key(
            [dv for s in specs for dv in s.DataVariables], key="DvID"
        )
        merged.Events = EquipmentExtractor._dedup_by_key(
            [ev for s in specs for ev in s.Events], key="CEID"
        )
        merged.Alarms = EquipmentExtractor._dedup_by_key(
            [al for s in specs for al in s.Alarms], key="AlarmID"
        )
        merged.RemoteCommands = EquipmentExtractor._dedup_by_key(
            [rc for s in specs for rc in s.RemoteCommands], key="RCMD"
        )
        merged.States = EquipmentExtractor._dedup_by_key(
            [st for s in specs for st in s.States], key="StateID"
        )
        merged.StateTransitions = EquipmentExtractor._dedup_transitions(
            [tr for s in specs for tr in s.StateTransitions]
        )
        
        # Merge Reports
        reports = []
        for s in specs:
            if hasattr(s, "Reports") and s.Reports:
                reports.extend(s.Reports)
        if reports:
            merged.Reports = EquipmentExtractor._dedup_by_key(reports, key="RPTID")

        return merged

    @staticmethod
    def _dedup_by_key(items: list, key: str) -> list:
        """Keep the highest-confidence entry per unique key value, merging list fields."""
        best: dict = {}
        for item in items:
            id_val = getattr(item, key)
            if id_val not in best:
                best[id_val] = item.model_copy(deep=True) if hasattr(item, "model_copy") else item.copy()
                continue

            existing = best[id_val]
            new_item = item
            
            conf_new = getattr(new_item, "Confidence", 1.0)
            conf_ext = getattr(existing, "Confidence", 1.0)

            if conf_new > conf_ext:
                existing, new_item = new_item.model_copy(deep=True) if hasattr(new_item, "model_copy") else new_item.copy(), existing

            if hasattr(existing, "model_fields"):
                for field in existing.model_fields.keys():
                    val_ext = getattr(existing, field)
                    val_new = getattr(new_item, field)

                    if isinstance(val_ext, list) and isinstance(val_new, list):
                        merged_list = []
                        for x in val_ext + val_new:
                            if x not in merged_list:
                                merged_list.append(x)
                        setattr(existing, field, merged_list)
                    elif isinstance(val_ext, str) and isinstance(val_new, str):
                        if val_ext in ("-", "", "unknown", "none") and val_new not in ("-", "", "unknown", "none"):
                            setattr(existing, field, val_new)
                        elif field == "Description" and len(val_new) > len(val_ext) and val_new not in ("-", "", "unknown", "none"):
                            setattr(existing, field, val_new)

            best[id_val] = existing

        return list(best.values())

    @staticmethod
    def _dedup_transitions(transitions: list) -> list:
        """Deduplicate state transitions by (FromState, ToState) tuple."""
        seen: dict[tuple[str, str], object] = {}
        for tr in transitions:
            key = (tr.FromState, tr.ToState)
            if key not in seen:
                seen[key] = tr
        return list(seen.values())

    # ------------------------------------------------------------------
    # Sanitization
    # ------------------------------------------------------------------

    @staticmethod
    def _sanitize(data: dict) -> None:
        """Clean up malformed transition entries before validation."""
        transitions = data.get("StateTransitions") or []
        data["StateTransitions"] = [
            t for t in transitions
            if t.get("FromState") and t.get("ToState")
        ]

    # ------------------------------------------------------------------
    # Prompt Builder
    # ------------------------------------------------------------------

    def _build_prompt(self, pdf_text: str) -> str:
        return f"""You extract semiconductor equipment SECS/GEM specs from technical documentation.

Output a single JSON object that conforms to this structure. No prose, no markdown fences, no commentary.
CRITICAL: DO NOT output the schema itself. You must output a JSON object populated with the actual extracted data.

EXPECTED JSON FORMAT:
{{
  "DocumentType": "string (User Manuals|Troubleshooting Guidance|GEM Manual|Variable Files|SML Scripts)",
  "ToolID": "string",
  "ToolType": "string",
  "Model": "string (optional)",
  "Protocol": "SECS/GEM",
  "StatusVariables":[
    {{
      "SVID": 123,
      "Name": "string",
      "Description": "string",
      "DataType": "string",
      "AccessType": "string",
      "Value": "string",
      "Confidence": 0.0 to 1.0 (double)
    }}
  ],
  "DataVariables": [
    {{
      "DvID": 123,
      "Name": "string",
      "ValueType": "string (float|integer|string|boolean)",
      "Unit": "string",
      "Confidence": 0.0 to 1.0 (double)
    }}
  ],
  "Events": [
    {{
      "CEID": 123,
      "Name": "string",
      "Description": "string",
      "LinkedVIDs": [123],
      "LinkedReports": ["string"],
      "Confidence": 0.0 to 1.0 (double)
    }}
  ],
  "Alarms": [
    {{
      "AlarmID": 123,
      "Name": "string",
      "Severity": "critical|warning|info",
      "LinkedVID": 123,
      "Description": "string",
      "Confidence": 0.0 to 1.0
    }}
  ],
  "RemoteCommands": [
    {{
      "RCMD": "string - the name of the command/message",
      "Description": "string",
      "Parameters": [{{ "Name": "string", "Type": "string" }}],
      "Confidence": 0.0 to 1.0
    }}
  ],
  "States": [
    {{ "StateID": "string", "Name": "string", "Description": "string" }}
  ],
  "StateTransitions": [
    {{
      "FromState": "string",
      "ToState": "string",
      "TriggerEvent": "string",
      "TriggerCommand": "string",
      "Manual": false
    }}
  ]
}}

CONFIDENCE RUBRIC (set the `Confidence` field per-entity):
- 1.0  = verbatim from a clearly labeled table
- 0.7  = inferred from prose with a clear section heading or strong signal
- <0.5 = guessed from weak context

HARD RULES:
- For every Event, you MUST extract the LinkedVIDs (Variables) associated with that event.
- Populate `LinkedVIDs` from any "Linked VIDs" column or comma-separated VID list following an event description.
- Populate `StateTransitions` from text matching `<State> -> <State> (Triggered by <event/command>)` patterns.
  - `FromState` and `ToState` are REQUIRED, must be non-null state names matching entries in `States`.
  - If the trigger is an event, set `TriggerEvent` (the event name) and leave `TriggerCommand` null.
  - If the trigger is a command, set `TriggerCommand` (the RCMD name) and leave `TriggerEvent` null.
  - If the transition is described as "manual" with no event/command, set `Manual: true` and leave both trigger fields null.
  - Never set both `TriggerEvent` and `TriggerCommand`.
- `Severity` MUST be lowercase: `critical`, `warning`, or `info`.
- CRITICAL: DO NOT hallucinate standard GEM capabilities (e.g., "Establish Communications", "Spooling", "Alarm Management") as Events or Alarms. ONLY extract entities that are explicitly defined with specific IDs in the document.
- If a section is absent from the document, return an empty list. Do NOT invent IDs, names, or fields.
- Output a single JSON object that validates against the schema above. Nothing else.

### DOCUMENT
{pdf_text}
"""

    # ------------------------------------------------------------------
    # Table Extraction (PDF → raw CSV → LLM)
    # ------------------------------------------------------------------

    _SHEET_NAME_KEYWORDS: dict[str, set[str]] = {
        "StatusVariables": {"svid", "sv", "status variable", "status_var", "status"},
        "DataVariables": {"dvid", "dv", "data variable", "data_var"},
        "Events": {"ceid", "ce", "collection event", "event"},
        "Alarms": {"alarm", "alid"},
        "RemoteCommands": {"rcmd", "remote command", "command"},
        "States": {"state"},
        "StateTransitions": {"transition", "state machine"},
        "Reports": {"report", "rptid", "reports"},
    }

    _SECTION_KEYWORDS: dict[str, set[str]] = {
        "StatusVariables": {
            "svid", "sv id", "status variable", "statusvariable",
            "variable id", "variable name", "access type",
        },
        "DataVariables": {
            "dvid", "dv id", "data variable", "datavariable",
            "value type", "valuetype",
        },
        "Events": {
            "ceid", "ce id", "collection event", "collectionevent",
            "event id", "event name", "linked vid",
        },
        "Alarms": {
            "alarm id", "alarmid", "alarm name", "alarm text",
            "alarm code", "alid", "al id", "severity",
        },
        "RemoteCommands": {
            "rcmd", "remote command", "remotecommand", "command name",
            "command id", "command",
        },
        "States": {
            "state id", "stateid", "state name", "equipment state",
        },
        "StateTransitions": {
            "from state", "to state", "fromstate", "tostate",
            "transition", "trigger",
        },
        "Reports": {
            "report id", "rptid", "rpt id", "report name",
        },
    }

    _CSV_FILENAMES: dict[str, str] = {
        "StatusVariables": "status_variables.csv",
        "DataVariables": "data_variables.csv",
        "Events": "events.csv",
        "Alarms": "alarms.csv",
        "RemoteCommands": "remote_commands.csv",
        "States": "states.csv",
        "StateTransitions": "state_transitions.csv",
        "Reports": "reports.csv",
    }

    _TABLE_PROMPTS: dict[str, str] = {
        "StatusVariables": """You are a SECS/GEM integration expert.
The table below was extracted directly from a semiconductor equipment manual.
The column names may differ from standard SECS/GEM field names — map them intelligently.
Convert every data row into a JSON array of StatusVariable objects.
Each object MUST have:
  "SVID": integer, "Name": string, "Description": string (empty if absent),
  "DataType": string (e.g. "U4","ASCII","Float","String"),
  "AccessType": string (e.g. "RO","RW"), "Value": string (empty if absent),
  "Confidence": float (1.0 if SVID+Name present else 0.7)
Return ONLY: {{"StatusVariables": [...]}}  No prose, no markdown fences.
TABLE (CSV):
{csv}""",

        "DataVariables": """You are a SECS/GEM integration expert.
The table below was extracted directly from a semiconductor equipment manual.
The column names may differ from standard SECS/GEM field names — map them intelligently.
Convert every data row into a JSON array of DataVariable objects.
Each object MUST have:
  "DvID": integer, "Name": string,
  "ValueType": string (float|integer|string|boolean),
  "Unit": string (empty if absent)
Return ONLY: {{"DataVariables": [...]}}  No prose, no markdown fences.
TABLE (CSV):
{csv}""",

        "Events": """You are a SECS/GEM integration expert.
The table below was extracted directly from a semiconductor equipment manual.
The column names may differ from standard SECS/GEM field names — map them intelligently.
Convert every data row into a JSON array of Collection Event objects.
CRITICAL: If a row does not contain an explicit numeric ID (CEID), IGNORE IT completely. Do NOT invent IDs for standard capabilities.
Each object MUST have:
  "CEID": integer, "Name": string, "Description": string (empty if absent),
  "LinkedVIDs": [integer] (parse comma-separated VID lists, empty list if absent),
  "LinkedReports": ["string"] (parse explicit report IDs if mentioned, else empty),
  "Confidence": float (1.0 if CEID+Name present else 0.7)
Return ONLY: {{"Events": [...]}}  No prose, no markdown fences.
TABLE (CSV):
{csv}""",

        "Alarms": """You are a SECS/GEM integration expert.
The table below was extracted directly from a semiconductor equipment manual.
The column names may differ from standard SECS/GEM field names — map them intelligently.
Convert every data row into a JSON array of Alarm objects.
CRITICAL: If a row does not contain an explicit numeric ID (AlarmID), IGNORE IT completely. Do NOT invent IDs.
Each object MUST have:
  "AlarmID": integer, "Name": string,
  "Severity": exactly one of "critical"|"warning"|"info" (lowercase),
  "LinkedVID": integer or null, "Description": string (empty if absent),
  "Confidence": float (1.0 if AlarmID+Name present else 0.7)
Return ONLY: {{"Alarms": [...]}}  No prose, no markdown fences.
TABLE (CSV):
{csv}""",

        "RemoteCommands": """You are a SECS/GEM integration expert.
The table below was extracted directly from a semiconductor equipment manual.
The column names may differ from standard SECS/GEM field names — map them intelligently.
Convert every data row into a JSON array of RemoteCommand objects.
Each object MUST have:
  "RCMD": string, "Description": string (empty if absent),
  "Parameters": [{{"Name": string, "Type": string}}] (parse from Parameters column, empty list if None/absent),
  "Confidence": float (1.0 if command name present else 0.7)
Return ONLY: {{"RemoteCommands": [...]}}  No prose, no markdown fences.
TABLE (CSV):
{csv}""",

        "States": """You are a SECS/GEM integration expert.
The table below was extracted directly from a semiconductor equipment manual.
Convert every data row into a JSON array of State objects.
Each object MUST have:
  "StateID": string, "Name": string, "Description": string (empty if absent)
Return ONLY: {{"States": [...]}}  No prose, no markdown fences.
TABLE (CSV):
{csv}""",

        "StateTransitions": """You are a SECS/GEM integration expert.
The table below was extracted directly from a semiconductor equipment manual.
The column names may differ — map them intelligently.
Convert every data row into a JSON array of StateTransition objects.
Each object MUST have:
  "FromState": string (REQUIRED), "ToState": string (REQUIRED),
  "TriggerEvent": string or null, "TriggerCommand": string or null,
  "Manual": boolean (true only if no event/command trigger)
Never set both TriggerEvent and TriggerCommand on the same entry.
Return ONLY: {{"StateTransitions": [...]}}  No prose, no markdown fences.
TABLE (CSV):
{csv}""",

        "Reports": """You are a SECS/GEM integration expert.
The table below was extracted directly from a semiconductor equipment manual.
The column names may differ from standard SECS/GEM field names — map them intelligently.
Convert every data row into a JSON array of ReportDefinition objects.
Each object MUST have:
  "RPTID": string (or integer), "Name": string,
  "LinkedVIDs": [integer] (parse comma-separated VID lists, empty list if absent)
Return ONLY: {{"Reports": [...]}}  No prose, no markdown fences.
TABLE (CSV):
{csv}""",
    }

    def _classify_table(self, rows: list[list[str]]) -> str | None:
        """Return the section name that best matches the table header, or None."""
        if not rows:
            return None
        header_text = " ".join(rows[0]).lower()
        best_section: str | None = None
        best_hits = 0
        for section, keywords in self._SECTION_KEYWORDS.items():
            hits = sum(1 for kw in keywords if kw in header_text)
            if hits > best_hits:
                best_hits = hits
                best_section = section
        return best_section if best_hits > 0 else None

    def _classify_by_sheet_name(self, sheet_name: str) -> str | None:
        """Return the section name that best matches an Excel sheet name, or None."""
        name_lower = sheet_name.lower()
        for section, keywords in self._SHEET_NAME_KEYWORDS.items():
            if any(kw in name_lower for kw in keywords):
                return section
        return None

    # ------------------------------------------------------------------
    # Excel Extraction
    # ------------------------------------------------------------------

    def extract_excel(self, excel_path: Union[str, Path]) -> EquipmentSpec:
        """Extract a SECS/GEM EquipmentSpec from an Excel workbook.

        Each sheet is converted to CSV and fed to the same per-section LLM
        prompts used for PDF table extraction. Sheets are classified first by
        their column headers, then by sheet name as fallback.
        """
        import openpyxl

        excel_path = Path(excel_path)
        try:
            wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
        except Exception as exc:
            logger.error("Failed to open Excel file %s: %s", excel_path.name, exc)
            return EquipmentSpec(ToolID="", ToolType="")

        partial_specs: list[EquipmentSpec] = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = [
                [str(cell.value).strip() if cell.value is not None else "" for cell in row]
                for row in ws.iter_rows()
            ]
            rows = [row for row in rows if any(cell for cell in row)]
            if len(rows) < 2:
                continue

            section = self._classify_table(rows) or self._classify_by_sheet_name(sheet_name)
            if not section:
                logger.info("Sheet '%s' could not be classified — skipping.", sheet_name)
                continue

            buf = io.StringIO()
            writer = csv.writer(buf)
            writer.writerows(rows)
            csv_str = buf.getvalue()

            logger.info("Extracting sheet '%s' as section '%s'", sheet_name, section)
            partial = self._extract_from_csv(section, csv_str)
            if partial:
                partial_specs.append(partial)

        wb.close()

        if not partial_specs:
            logger.warning("No classifiable sheets found in %s", excel_path.name)
            return EquipmentSpec(ToolID="", ToolType="")

        merged = self._merge_specs(partial_specs)
        logger.info(
            "Excel extraction done: %d SVs, %d DVs, %d Events, %d Alarms",
            len(merged.StatusVariables), len(merged.DataVariables),
            len(merged.Events), len(merged.Alarms),
        )
        return merged

    def extract_and_save_tables(
        self,
        pdf_path: Path,
        tables_dir: Union[Path, None],
       
        tables_store_path: Union[Path, None] = None,
    ) -> dict[str, str]:
        """
        Extract all tables from the PDF using pdfplumber.
        Classify each table by section.
        Save each classified table as-is to a CSV file (raw column names preserved).
        Return a dict of {section: csv_string} for LLM consumption.
        """
        section_rows: dict[str, list[list[str]]] = {}

        try:
            with pdfplumber.open(pdf_path) as pdf:
                for page in pdf.pages:
                    for raw_table in (page.extract_tables() or []):
                        cleaned = [
                            [str(cell).strip() if cell is not None else "" for cell in row]
                            for row in raw_table
                            if any(cell not in [None, "", " "] for cell in row)
                        ]
                        if len(cleaned) < 2:
                            continue
                        section = self._classify_table(cleaned)
                        if section:
                            if section not in section_rows:
                                section_rows[section] = cleaned
                            else:
                                # Append data rows only (skip duplicate header)
                                section_rows[section].extend(cleaned[1:])
        except Exception as exc:
            logger.error("pdfplumber failed on %s: %s", pdf_path.name, exc)
            return {}

        if not section_rows:
            logger.info("No classifiable tables found in %s", pdf_path.name)
            return {}

        section_csvs: dict[str, str] = {}

        for section, rows in section_rows.items():
            # Serialise to CSV string
            buf = io.StringIO()
            writer = csv.writer(buf)
            writer.writerows(rows)
            csv_str = buf.getvalue()
            section_csvs[section] = csv_str

            # Save to disk if tables_dir provided
            if tables_dir is not None:
                raw_tables_dir = tables_dir / "RawTables"
                raw_tables_dir.mkdir(parents=True, exist_ok=True)
                csv_path = raw_tables_dir / self._CSV_FILENAMES[section]
                # Append new data rows to existing file, or create fresh
                if csv_path.exists():
                    existing_lines = csv_path.read_text(encoding="utf-8").splitlines()
                    new_lines = csv_str.splitlines()
                    # Skip header of incoming if file already has content
                    combined = existing_lines + new_lines[1:] if existing_lines else new_lines
                    csv_path.write_text("\n".join(combined) + "\n", encoding="utf-8")
                else:
                    csv_path.write_text(csv_str, encoding="utf-8")
                logger.info("Saved raw %s table (%d rows) to %s", section, len(rows), csv_path)

        # ── Index table rows into the dedicated 'tables' vector store ──────────
        if tables_store_path is not None and section_rows:
            try:
                from source.utils.embedder import VectorStoreManager
                tables_vs = VectorStoreManager(tables_store_path)
                for section, rows in section_rows.items():
                    if len(rows) < 2:
                        continue
                    headers = rows[0]
                    for row_idx, row in enumerate(rows[1:], start=1):
                        parts = []
                        for header, cell in zip(headers, row):
                            if cell:
                                parts.append(f"{header}: {cell}")
                        if not parts:
                            continue
                        sentence = f"[{section}] " + " | ".join(parts)
                        tables_vs.add_document(
                            sentence,
                            metadata={
                                "project_id": str(pdf_path.parent.parent.name),
                                "document_id": pdf_path.stem,
                                "document_category": "tables",
                                "section": section,
                                "row_index": row_idx,
                            },
                        )
                logger.info(
                    "Indexed table rows from %s into tables vector store at %s",
                    pdf_path.name, tables_store_path,
                )
            except Exception as exc:
                logger.warning(
                    "Table vector store indexing failed for %s (non-fatal): %s",
                    pdf_path.name, exc,
                )

        return section_csvs

    def _extract_from_csv(self, section: str, csv_str: str) -> EquipmentSpec | None:
        """Send a raw CSV table to the LLM with a section-specific prompt."""
        prompt_template = self._TABLE_PROMPTS.get(section)
        if not prompt_template:
            return None

        prompt = prompt_template.format(csv=csv_str)

        for attempt, llm in enumerate((self._llm, self._llm_retry), start=1):
            try:
                raw = llm.invoke(prompt).content
                if isinstance(raw, list):
                    raw = "".join(
                        p.get("text", "") if isinstance(p, dict) else str(p) for p in raw
                    )
                return self._parse_table_response(raw, section)
            except Exception as exc:
                logger.warning(
                    "Table LLM attempt %d failed for section %s: %s", attempt, section, exc
                )

        logger.error("Both LLM attempts failed for section %s — skipping.", section)
        return None

    def _parse_table_response(self, raw: str, section: str) -> EquipmentSpec | None:
        """Parse LLM JSON response for a single table section into a partial EquipmentSpec."""
        try:
            data = json.loads(raw)
        except json.JSONDecodeError as exc:
            logger.warning("JSON parse failed for %s table response: %s", section, exc)
            return None

        spec = EquipmentSpec(ToolID="", ToolType="")
        try:
            items = data.get(section) or []
            if section == "StatusVariables":
                spec.StatusVariables = [StatusVariable.model_validate(i) for i in items]
            elif section == "DataVariables":
                spec.DataVariables = [DataVariable.model_validate(i) for i in items]
            elif section == "Events":
                spec.Events = [Event.model_validate(i) for i in items]
            elif section == "Alarms":
                spec.Alarms = [Alarm.model_validate(i) for i in items]
            elif section == "RemoteCommands":
                spec.RemoteCommands = [RemoteCommand.model_validate(i) for i in items]
            elif section == "States":
                spec.States = [State.model_validate(i) for i in items]
            elif section == "StateTransitions":
                valid = [t for t in items if t.get("FromState") and t.get("ToState")]
                spec.StateTransitions = [StateTransition.model_validate(t) for t in valid]
            elif section == "Reports":
                from source.schemas.secsgem import ReportDefinition
                spec.Reports = [ReportDefinition.model_validate(i) for i in items]
        except Exception as exc:
            logger.warning("model_validate failed for %s table response: %s", section, exc)
            return None

        return spec
