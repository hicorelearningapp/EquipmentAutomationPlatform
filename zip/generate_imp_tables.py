"""
generate_imp_tables.py
======================
Generates the companion Excel workbook with 24 styled worksheets containing all
SECS/GEM tabular data for the IMP-HE2000 High-Energy Ion Implanter.
"""

import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_sheet(wb, title, headers, rows):
    # Retrieve or create sheet
    if title in wb.sheetnames:
        ws = wb[title]
        # Clear existing contents
        ws.delete_rows(1, ws.max_row + 1)
    else:
        ws = wb.create_sheet(title=title)
    
    # Write headers
    ws.append(headers)
    
    # Write rows
    for r in rows:
        ws.append(r)
        
    # Styles definition
    header_fill = PatternFill(start_color="1B3A6B", end_color="1B3A6B", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    
    alt_fill = PatternFill(start_color="F4F6FA", end_color="F4F6FA", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    border_side = Side(border_style="thin", color="B0BDD6")
    cell_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    
    # Header styling
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = cell_border
        
    # Data row styling
    for row_idx in range(2, ws.max_row + 1):
        row_fill = alt_fill if row_idx % 2 == 0 else white_fill
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.fill = row_fill
            cell.font = Font(name="Calibri", size=10)
            cell.border = cell_border
            
            # Alignments
            h_text = str(headers[col_idx-1]).lower()
            if any(x in h_text for x in ["id", "code", "type", "unit", "range", "update", "direction", "stream", "status", "from state", "to state", "trigger", "manual"]):
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

    # Freeze panes (row 1 frozen)
    ws.freeze_panes = "A2"
    
    # AutoFilter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
    
    # Auto-fit columns
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val = str(cell.value or "")
            if len(val) > max_len:
                max_len = len(val)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

def generate_workbook(output_path):
    wb = openpyxl.Workbook()
    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)

    # 1. Tool Identity
    tool_identity_headers = ["Field", "Value"]
    tool_identity_rows = [
        ["Tool ID", "IMP-HE2000-FAB002"],
        ["Tool Type", "High-Energy Medium-Current Ion Implanter"],
        ["Model", "IMP-HE2000"],
        ["Manufacturer", "PrecisionBeam Technologies"],
        ["Serial Number", "PB-IMP-2025-0017"],
        ["Firmware Version", "v6.2.1"],
        ["GEM Version", "SEMI E30-0618"],
        ["Installation Site", "Fab-Beta / Bay 3 / Row A"],
        ["Wafer Diameter", "300 mm"],
        ["Process Types", "Boron, Phosphorus, Arsenic, BF₂ Implantation"],
        ["Energy Range", "10 – 200 keV"],
        ["Current Range", "1 µA – 20 mA"],
    ]
    create_sheet(wb, "Tool Identity", tool_identity_headers, tool_identity_rows)

    # 2. HSMS Parameters
    hsms_headers = ["Parameter", "Value", "Notes"]
    hsms_rows = [
        ["Transport", "HSMS-SS", "Single Session"],
        ["Mode", "Passive", "Equipment listens; host connects"],
        ["IP Address", "10.30.14.35", "Static — VLAN 30"],
        ["Port", "5020", "GEM port assignment"],
        ["T3 Reply Timeout", "45 s", "Per SEMI E37"],
        ["T5 Connect Timeout", "10 s", "Per SEMI E37"],
        ["T6 Control Timeout", "5 s", "Per SEMI E37"],
        ["T7 Not Selected", "10 s", "Per SEMI E37"],
        ["T8 Network Timeout", "5 s", "Per SEMI E37"],
        ["DeviceID", "1", "Single logical device"],
        ["Max Open Transactions", "127", "Simultaneous S/F pairs"],
    ]
    create_sheet(wb, "HSMS Parameters", hsms_headers, hsms_rows)

    # 3. VID Range Map
    vid_map_headers = ["VID Range", "Subsystem", "Description"]
    vid_map_rows = [
        ["1000 – 1049", "Global / System", "Tool-wide status and control variables"],
        ["1050 – 1099", "Ion Source", "Arc current, filament, gas flow, extraction"],
        ["1100 – 1149", "Beamline", "Analyzer magnet, beam current, energy, scanning"],
        ["1150 – 1199", "End Station", "Platen tilt/twist, dose, temperature, Faraday"],
        ["1200 – 1249", "Wafer Handler", "Robot state, load port, load lock, pre-aligner"],
        ["1250 – 1299", "Vacuum System", "Beamline vacuum, end station vacuum, load lock"],
        ["1300 – 1349", "High Voltage & Power", "Terminal HV, extraction, suppression, PSU status"],
        ["2000 – 2049", "Data Variables (Recipe)", "Recipe parameters and implant targets"],
        ["2050 – 2099", "Data Variables (Results)", "Process results and metrology"],
        ["3000 – 3099", "Collection Events", "Process and system events"],
        ["3100 – 3199", "Reports (RPTIDs)", "Named variable sets for event linking"],
        ["4000 – 4099", "Alarms", "Equipment alarm codes"],
        ["5000 – 5049", "Equipment Constants", "User-configurable setpoints"],
    ]
    create_sheet(wb, "VID Range Map", vid_map_headers, vid_map_rows)

    # SV Headers
    sv_headers = ["SVID", "Name", "Type", "Unit", "Range", "Update Rate", "Description"]

    # 4. SV — Global System
    sv_global_rows = [
        [1001, "SystemStatus", "STRING", "—", "IDLE, IMPLANTING, TUNING, FAULT, MAINTENANCE", "1 s", "Overall status: IDLE, IMPLANTING, TUNING, FAULT, MAINTENANCE"],
        [1002, "ControlState", "U1", "—", "1–5", "1 s", "GEM control state: 1=OFF-LINE, 2=ATTEMPT ONLINE, 3=HOST, 4=EQUIPMENT, 5=ON-LINE"],
        [1003, "ProcessState", "STRING", "—", "See Ch.10", "1 s", "Current equipment process state from the state machine"],
        [1004, "PreviousProcessState", "STRING", "—", "See Ch.10", "1 s", "Previous process state (set at each transition)"],
        [1005, "EquipmentTime", "STRING", "—", "ISO 8601", "1 s", "Equipment real-time clock YYYYMMDDHHmmss"],
        [1006, "ActiveWaferCount", "U2", "—", "0–25", "1 s", "Total wafers currently inside the tool"],
        [1007, "EMOStatus", "BOOLEAN", "—", "0/1", "1 s", "TRUE = Emergency Master Off activated"],
        [1008, "MaintenanceMode", "BOOLEAN", "—", "0/1", "1 s", "TRUE = maintenance mode active, processing inhibited"],
        [1009, "RadiationInterlock", "BOOLEAN", "—", "0/1", "1 s", "TRUE = X-ray radiation level exceeds safe threshold"],
        [1010, "DoorInterlock_Source", "BOOLEAN", "—", "0/1", "500 ms", "TRUE = ion source access panel is open"],
        [1011, "DoorInterlock_Beamline", "BOOLEAN", "—", "0/1", "500 ms", "TRUE = beamline access panel is open"],
        [1012, "DoorInterlock_EndStation", "BOOLEAN", "—", "0/1", "500 ms", "TRUE = end station access door is open"],
        [1013, "PMCounter_Source", "U4", "hours", "0–9999", "5 s", "Ion source operating hours since last filament replacement"],
        [1014, "PMCounter_Beamline", "U4", "hours", "0–9999", "5 s", "Beamline operating hours since last PM"],
    ]
    create_sheet(wb, "SV - Global System", sv_headers, sv_global_rows)

    # 5. SV — Ion Source
    sv_source_rows = [
        [1051, "Source_ArcCurrent", "FLOAT", "A", "0–5.0", "500 ms", "Ion source arc discharge current"],
        [1052, "Source_ArcVoltage", "FLOAT", "V", "0–120", "500 ms", "Arc discharge voltage across the source chamber"],
        [1053, "Source_ArcCurrentSetpoint", "FLOAT", "A", "0–5.0", "2 s", "Active recipe arc current setpoint (R/W)"],
        [1054, "Source_FilamentCurrent", "FLOAT", "A", "0–200", "500 ms", "Filament heating current"],
        [1055, "Source_FilamentVoltage", "FLOAT", "V", "0–15", "500 ms", "Filament heating voltage"],
        [1056, "Source_GasFlow", "FLOAT", "sccm", "0–10", "500 ms", "MFC actual flow rate for source gas"],
        [1057, "Source_GasFlowSetpoint", "FLOAT", "sccm", "0–10", "2 s", "MFC flow setpoint (R/W)"],
        [1058, "Source_GasSpecies", "STRING", "—", "BF3, PH3, AsH3, Ar, Xe", "5 s", "Active source gas: BF3, PH3, AsH3, Ar, Xe"],
        [1059, "Source_ExtractionVoltage", "FLOAT", "kV", "0–40", "500 ms", "Extraction electrode voltage"],
        [1060, "Source_ExtractionCurrent", "FLOAT", "mA", "0–50", "500 ms", "Extraction beam current at source exit"],
        [1061, "Source_LifetimeHours", "FLOAT", "hours", "0–500", "5 s", "Cumulative source operating hours (resets at filament change)"],
        [1062, "Source_Status", "STRING", "—", "OFF, WARMING_UP, PLASMA_ON, STABLE, FAULT", "1 s", "Source state: OFF, WARMING_UP, PLASMA_ON, STABLE, FAULT"],
        [1063, "Source_VaporiserTemp", "FLOAT", "°C", "0–300", "2 s", "Vaporiser temperature for solid source materials (e.g., Decaborane)"],
        [1064, "Source_ChamberPressure", "FLOAT", "mTorr", "0–500", "500 ms", "Ion source chamber internal pressure"],
    ]
    create_sheet(wb, "SV - Ion Source", sv_headers, sv_source_rows)

    # 6. SV — Beamline
    sv_beamline_rows = [
        [1101, "Beam_AnalyzerMagnetCurrent", "FLOAT", "A", "0–300", "500 ms", "Analyzer magnet coil current"],
        [1102, "Beam_AnalyzerMagnetSetpoint", "FLOAT", "A", "0–300", "2 s", "Analyzer magnet current setpoint for selected species/energy"],
        [1103, "Beam_ResolvedCurrent", "FLOAT", "mA", "0–30", "500 ms", "Beam current at the resolving aperture Faraday cup"],
        [1104, "Beam_Energy", "FLOAT", "keV", "0–200", "500 ms", "Actual measured beam energy"],
        [1105, "Beam_EnergySetpoint", "FLOAT", "keV", "0–200", "2 s", "Recipe-specified beam energy setpoint"],
        [1106, "Beam_AccelVoltage", "FLOAT", "kV", "0–200", "500 ms", "Acceleration column voltage"],
        [1107, "Beam_DecelVoltage", "FLOAT", "kV", "0–50", "500 ms", "Deceleration electrode voltage (for energy contamination control)"],
        [1108, "Beam_ScanFreqX", "FLOAT", "Hz", "0–1500", "2 s", "X-axis electrostatic scan frequency"],
        [1109, "Beam_ScanFreqY", "FLOAT", "Hz", "0–100", "2 s", "Y-axis mechanical scan frequency"],
        [1110, "Beam_ScanAmplitudeX", "FLOAT", "mm", "0–350", "2 s", "X-axis scan amplitude at wafer plane"],
        [1111, "Beam_ScanAmplitudeY", "FLOAT", "mm", "0–350", "2 s", "Y-axis scan amplitude at wafer plane"],
        [1112, "Beam_SpotSizeX", "FLOAT", "mm", "0–50", "5 s", "Beam spot width at wafer plane (X)"],
        [1113, "Beam_SpotSizeY", "FLOAT", "mm", "0–50", "5 s", "Beam spot height at wafer plane (Y)"],
        [1114, "Beam_Parallelism", "FLOAT", "deg", "0–2.0", "5 s", "Beam angle deviation from parallel (should be < 0.5°)"],
        [1115, "Beam_Status", "STRING", "—", "OFF, TUNING, BEAM_ON, STABLE, FAULT", "1 s", "Beamline state: OFF, TUNING, BEAM_ON, STABLE, FAULT"],
    ]
    create_sheet(wb, "SV - Beamline", sv_headers, sv_beamline_rows)

    # 7. SV — End Station
    sv_endstation_rows = [
        [1151, "EndStn_PlatenTilt", "FLOAT", "deg", "0–60", "500 ms", "Wafer platen tilt angle (relative to beam)"],
        [1152, "EndStn_PlatenTiltSetpoint", "FLOAT", "deg", "0–60", "2 s", "Recipe tilt angle setpoint"],
        [1153, "EndStn_PlatenTwist", "FLOAT", "deg", "0–360", "500 ms", "Wafer platen twist angle"],
        [1154, "EndStn_PlatenTwistSetpoint", "FLOAT", "deg", "0–360", "2 s", "Recipe twist angle setpoint"],
        [1155, "EndStn_DoseCurrent", "FLOAT", "µA", "0–20000", "200 ms", "Instantaneous dose Faraday cup current"],
        [1156, "EndStn_DoseAccumulated", "FLOAT", "ions/cm²", "0–1e18", "200 ms", "Accumulated dose delivered to current wafer"],
        [1157, "EndStn_DoseTarget", "FLOAT", "ions/cm²", "0–1e18", "2 s", "Recipe target dose"],
        [1158, "EndStn_DoseUniformity", "FLOAT", "%", "0–10", "5 s", "Real-time dose uniformity (1-sigma)"],
        [1159, "EndStn_WaferTemperature", "FLOAT", "°C", "0–300", "1 s", "Wafer surface temperature (IR pyrometer)"],
        [1160, "EndStn_WaferPresent", "BOOLEAN", "—", "0/1", "500 ms", "TRUE = wafer detected on platen"],
        [1161, "EndStn_WaferClamped", "BOOLEAN", "—", "0/1", "500 ms", "TRUE = electrostatic clamp engaged"],
        [1162, "EndStn_CoolantFlow", "FLOAT", "L/min", "0–10", "2 s", "Platen backside He coolant flow rate"],
        [1163, "EndStn_CoolantPressure", "FLOAT", "Torr", "0–20", "2 s", "Backside He coolant pressure"],
        [1164, "EndStn_ImplantTime", "FLOAT", "sec", "0–3600", "1 s", "Elapsed implant time for current wafer"],
        [1165, "EndStn_ProcessState", "STRING", "—", "IDLE, LOADING, ALIGNING, IMPLANTING, UNLOADING, FAULT", "1 s", "End station sub-state: IDLE, LOADING, ALIGNING, IMPLANTING, UNLOADING, FAULT"],
    ]
    create_sheet(wb, "SV - End Station", sv_headers, sv_endstation_rows)

    # 8. SV — Wafer Handler
    sv_handler_rows = [
        [1201, "WH_RobotStatus", "STRING", "—", "IDLE, TRANSFERRING, ALIGNING, FAULT, TEACHING", "1 s", "Robot state: IDLE, TRANSFERRING, ALIGNING, FAULT, TEACHING"],
        [1202, "WH_RobotPosition", "STRING", "—", "HOME, LOADPORT1, LOADPORT2, LOADLOCK, PREALIGNER, ENDSTATION", "500 ms", "Robot position: HOME, LOADPORT1, LOADPORT2, LOADLOCK, PREALIGNER, ENDSTATION"],
        [1203, "WH_LoadPort1_Status", "STRING", "—", "EMPTY, FOUP_PRESENT, MAPPING, READY, FAULT", "500 ms", "Port 1: EMPTY, FOUP_PRESENT, MAPPING, READY, FAULT"],
        [1204, "WH_LoadPort2_Status", "STRING", "—", "EMPTY, FOUP_PRESENT, MAPPING, READY, FAULT", "500 ms", "Port 2: EMPTY, FOUP_PRESENT, MAPPING, READY, FAULT"],
        [1205, "WH_LoadLock_State", "STRING", "—", "VENTED, PUMPING, VACUUM, VENTING, FAULT", "1 s", "Load lock: VENTED, PUMPING, VACUUM, VENTING, FAULT"],
        [1206, "WH_LoadLock_Pressure", "FLOAT", "Torr", "0–760", "500 ms", "Load lock chamber pressure"],
        [1207, "WH_PrealignAngle", "FLOAT", "deg", "0–360", "2 s", "Notch finder measured wafer orientation angle"],
        [1208, "WH_WaferInLoadLock", "BOOLEAN", "—", "0/1", "500 ms", "TRUE = wafer present in load lock"],
        [1209, "WH_FOUP1_WaferCount", "U1", "—", "0–25", "5 s", "Number of wafers detected in FOUP on port 1"],
        [1210, "WH_FOUP2_WaferCount", "U1", "—", "0–25", "5 s", "Number of wafers detected in FOUP on port 2"],
    ]
    create_sheet(wb, "SV - Wafer Handler", sv_headers, sv_handler_rows)

    # 9. SV — Vacuum System
    sv_vacuum_rows = [
        [1251, "Vac_Beamline_Pressure", "FLOAT", "Torr", "0–1e-3", "500 ms", "Beamline section vacuum pressure (ion gauge)"],
        [1252, "Vac_EndStation_Pressure", "FLOAT", "Torr", "0–1e-3", "500 ms", "End station process chamber vacuum pressure"],
        [1253, "Vac_SourceChamber_Pressure", "FLOAT", "Torr", "0–1", "500 ms", "Ion source region vacuum pressure"],
        [1254, "Vac_TurboPump1_Speed", "FLOAT", "RPM", "0–60000", "2 s", "Beamline turbomolecular pump rotor speed"],
        [1255, "Vac_TurboPump1_Status", "STRING", "—", "OFF, STARTING, RUNNING, FAULT", "2 s", "OFF, STARTING, RUNNING, FAULT"],
        [1256, "Vac_TurboPump2_Speed", "FLOAT", "RPM", "0–60000", "2 s", "End station turbomolecular pump rotor speed"],
        [1257, "Vac_TurboPump2_Status", "STRING", "—", "OFF, STARTING, RUNNING, FAULT", "2 s", "OFF, STARTING, RUNNING, FAULT"],
        [1258, "Vac_CryoPump_Temp", "FLOAT", "K", "0–300", "5 s", "Cryopump cold head temperature (target < 15 K)"],
        [1259, "Vac_CryoPump_Status", "STRING", "—", "OFF, COOLDOWN, RUNNING, REGENERATING, FAULT", "5 s", "OFF, COOLDOWN, RUNNING, REGENERATING, FAULT"],
        [1260, "Vac_RoughingPump_Status", "STRING", "—", "OFF, RUNNING, FAULT", "2 s", "Dry scroll pump: OFF, RUNNING, FAULT"],
        [1261, "Vac_GateValve_BL", "BOOLEAN", "—", "0/1", "500 ms", "TRUE = beamline gate valve OPEN"],
        [1262, "Vac_GateValve_ES", "BOOLEAN", "—", "0/1", "500 ms", "TRUE = end station gate valve OPEN"],
    ]
    create_sheet(wb, "SV - Vacuum System", sv_headers, sv_vacuum_rows)

    # 10. SV — High Voltage & Power
    sv_hv_rows = [
        [1301, "HV_TerminalVoltage", "FLOAT", "kV", "0–200", "500 ms", "Terminal high-voltage supply output"],
        [1302, "HV_TerminalCurrent", "FLOAT", "mA", "0–50", "500 ms", "Terminal high-voltage supply current"],
        [1303, "HV_TerminalSetpoint", "FLOAT", "kV", "0–200", "2 s", "Terminal voltage setpoint"],
        [1304, "HV_TerminalStatus", "STRING", "—", "OFF, RAMPING, ON, FAULT", "1 s", "OFF, RAMPING, ON, FAULT"],
        [1305, "HV_SuppressionVoltage", "FLOAT", "kV", "0–10", "500 ms", "Electron suppression electrode voltage"],
        [1306, "HV_ExtractionVoltage", "FLOAT", "kV", "0–40", "500 ms", "Extraction electrode voltage (same as Source_ExtractionVoltage)"],
        [1307, "HV_InterlockStatus", "BOOLEAN", "—", "0/1", "500 ms", "TRUE = all HV interlocks satisfied (safe to energise)"],
        [1308, "PSU_24V_Status", "BOOLEAN", "—", "0/1", "2 s", "TRUE = 24V instrument power supply OK"],
        [1309, "PSU_48V_Status", "BOOLEAN", "—", "0/1", "2 s", "TRUE = 48V servo power supply OK"],
        [1310, "PSU_MainAC_Status", "BOOLEAN", "—", "0/1", "2 s", "TRUE = main 3-phase AC power within tolerance"],
    ]
    create_sheet(wb, "SV - High Voltage and Power", sv_headers, sv_hv_rows)

    # DV Headers
    dv_headers = ["DVID", "Name", "Type", "Unit", "Description"]

    # 11. DV — Recipe & Implant Targets
    dv_recipe_rows = [
        [2001, "RecipeID", "STRING", "—", "Active process recipe identifier"],
        [2002, "RecipeVersion", "STRING", "—", "Recipe version string"],
        [2003, "LotID", "STRING", "—", "Lot identifier for the current batch"],
        [2004, "WaferID", "STRING", "—", "Wafer ID of wafer currently being processed"],
        [2005, "SlotNumber", "U1", "—", "FOUP slot number of current wafer (1–25)"],
        [2006, "TargetSpecies", "STRING", "—", "Dopant species: B, B11, BF2, P, As, Ar"],
        [2007, "TargetEnergy", "FLOAT", "keV", "Recipe-specified beam energy"],
        [2008, "TargetDose", "FLOAT", "ions/cm²", "Recipe-specified implant dose"],
        [2009, "TargetTilt", "FLOAT", "deg", "Recipe-specified wafer tilt angle"],
        [2010, "TargetTwist", "FLOAT", "deg", "Recipe-specified wafer twist angle"],
        [2011, "TargetBeamCurrent", "FLOAT", "mA", "Recipe-specified beam current setpoint"],
        [2012, "TargetDoseUniformity", "FLOAT", "%", "Maximum acceptable dose non-uniformity"],
        [2013, "TargetWaferTemp_Max", "FLOAT", "°C", "Maximum allowable wafer temperature during implant"],
        [2014, "ImplantMode", "STRING", "—", "SPOT_BEAM, RIBBON_BEAM, ENERGY_PURITY"],
    ]
    create_sheet(wb, "DV - Recipe and Implant Targets", dv_headers, dv_recipe_rows)

    # 12. DV — Process Results
    dv_results_rows = [
        [2051, "ActualDoseDelivered", "FLOAT", "ions/cm²", "Total measured dose delivered to the wafer"],
        [2052, "DoseUniformity_Measured", "FLOAT", "%", "Measured 1-sigma dose uniformity across the wafer"],
        [2053, "ActualEnergy", "FLOAT", "keV", "Average measured beam energy during implant"],
        [2054, "ActualBeamCurrent_Avg", "FLOAT", "mA", "Average beam current during implant"],
        [2055, "ActualTilt", "FLOAT", "deg", "Actual platen tilt angle during implant"],
        [2056, "ActualTwist", "FLOAT", "deg", "Actual platen twist angle during implant"],
        [2057, "ImplantDuration", "FLOAT", "sec", "Total implant time for this wafer"],
        [2058, "WaferTemp_Max", "FLOAT", "°C", "Peak wafer temperature recorded during implant"],
        [2059, "WaferTemp_Avg", "FLOAT", "°C", "Average wafer temperature during implant"],
        [2060, "ProcessResult", "STRING", "—", "PASS, FAIL, ABORTED, INCOMPLETE"],
        [2061, "AbortReason", "STRING", "—", "Plain-text reason if ProcessResult = ABORTED"],
        [2062, "BeamStability", "FLOAT", "%", "Beam current stability (σ/mean × 100) during implant"],
        [2063, "DoseAccuracy", "FLOAT", "%", "(ActualDose - TargetDose) / TargetDose × 100"],
        [2064, "EnergyContamination", "FLOAT", "%", "Fraction of beam at wrong energy (decel column leakage)"],
        [2065, "ParticleCount", "U4", "—", "Post-implant particle count on wafer (if equipped)"],
    ]
    create_sheet(wb, "DV - Process Results", dv_headers, dv_results_rows)

    # 13. Collection Events
    ce_headers = ["CEID", "Event Name", "Linked VIDs", "Trigger Condition"]
    ce_rows = [
        [3001, "ImplantStart", "1151,1155,2001,2004,2007,2008", "Beam turned on and scanning wafer"],
        [3002, "ImplantEnd", "2051,2052,2053,2054,2060", "Target dose reached or process completed"],
        [3003, "ImplantAbort", "2061,1003,1165", "Process aborted by host, operator, or fault"],
        [3004, "BeamTuneStart", "1051,1101,1104", "Beam tuning/setup sequence initiated"],
        [3005, "BeamTuneComplete", "1103,1104,1114,1115", "Beam stable and within specification"],
        [3006, "BeamLost", "1103,1115", "Beam current dropped below minimum threshold"],
        [3007, "SourceIgnition", "1051,1052,1062", "Ion source plasma successfully ignited"],
        [3008, "SourceExtinguish", "1062", "Ion source plasma turned off"],
        [3009, "SourceFault", "1051,1052,1061,1062", "Ion source fault detected"],
        [3010, "WaferLoaded_EndStation", "1160,1161,2004", "Wafer placed on end station platen"],
        [3011, "WaferUnloaded_EndStation", "1160,2004,2060", "Wafer removed from end station platen"],
        [3012, "WaferPrealigned", "1207,2004", "Notch finder alignment completed"],
        [3013, "FOUPLoaded_Port1", "1203,1209", "FOUP placed and mapped on Load Port 1"],
        [3014, "FOUPUnloaded_Port1", "1203", "FOUP removed from Load Port 1"],
        [3015, "FOUPLoaded_Port2", "1204,1210", "FOUP placed and mapped on Load Port 2"],
        [3016, "FOUPUnloaded_Port2", "1204", "FOUP removed from Load Port 2"],
        [3017, "LoadLockPumpDown", "1205,1206", "Load lock pump-down started"],
        [3018, "LoadLockVacuumReached", "1205,1206", "Load lock reached transfer vacuum"],
        [3019, "LoadLockVent", "1205,1206", "Load lock vent to atmosphere started"],
        [3020, "LoadLockVented", "1205,1206", "Load lock at atmospheric pressure"],
        [3021, "RobotFault", "1201,1202", "Wafer handler robot error"],
        [3022, "SlotMapComplete_Port1", "1203,1209", "FOUP slot mapping completed on Port 1"],
        [3023, "SlotMapComplete_Port2", "1204,1210", "FOUP slot mapping completed on Port 2"],
        [3030, "ControlStateChange", "1002", "GEM control state changed"],
        [3031, "RecipeLoaded", "2001,2002", "New recipe loaded into active slot"],
        [3032, "RecipeDeleted", "2001", "Recipe deleted from controller"],
        [3033, "MaintenanceModeEntered", "1008", "Equipment entered maintenance mode"],
        [3034, "MaintenanceModeExited", "1008", "Equipment exited maintenance mode"],
        [3035, "HV_Enabled", "1301,1304", "High voltage turned on"],
        [3036, "HV_Disabled", "1304", "High voltage turned off"],
        [3037, "AlarmSet", "1001", "Any alarm became active"],
        [3038, "AlarmCleared", "1001", "Any alarm was cleared"],
        [3039, "EquipmentClockSync", "1005", "Clock synchronized with host"],
        [3040, "PMDue_Source", "1013", "Source PM counter reached threshold"],
        [3041, "PMDue_Beamline", "1014", "Beamline PM counter reached threshold"],
    ]
    create_sheet(wb, "Collection Events", ce_headers, ce_rows)

    # 14. Reports
    report_headers = ["RPTID", "Report Name", "Linked VIDs", "Description"]
    report_rows = [
        [3101, "RPT_ImplantStart", "1051,1104,1151,1155,2001,2004,2006,2007,2008,2009,2010", "Source, beam, tilt, dose target at implant start"],
        [3102, "RPT_ImplantEnd", "2051,2052,2053,2054,2055,2056,2057,2058,2060,2062,2063", "All process results at implant end"],
        [3103, "RPT_BeamStatus", "1051,1052,1101,1103,1104,1108,1112,1113,1114,1115", "Complete beam health snapshot"],
        [3104, "RPT_WaferLoad", "1160,1161,2004,2005,1201,1005", "Wafer ID, slot, robot state at load"],
        [3105, "RPT_WaferUnload", "1160,2004,2060,1005", "Wafer ID, result, timestamp at unload"],
        [3106, "RPT_AlarmEvent", "1001,1005", "System status and timestamp at alarm"],
        [3107, "RPT_SourceStatus", "1051,1052,1054,1056,1058,1061,1062", "Full ion source health snapshot"],
        [3108, "RPT_VacuumStatus", "1251,1252,1253,1255,1257,1259", "All vacuum pressures and pump statuses"],
        [3109, "RPT_HVStatus", "1301,1302,1304,1305,1306,1307", "High voltage system status"],
        [3110, "RPT_ControlStateChange", "1002,1005", "Control state and timestamp"],
        [3111, "RPT_ProcessAbort", "2061,1003,1165,1005", "Abort reason, states, and timestamp"],
        [3112, "RPT_MaintenanceEvent", "1008,1013,1014,1005", "Maintenance status and PM counters"],
        [3113, "RPT_RecipeManagement", "2001,2002,1005", "Recipe ID, version, timestamp"],
        [3114, "RPT_DoseProfile", "1155,1156,1157,1158,2051,2052", "Real-time and final dose metrics"],
    ]
    create_sheet(wb, "Reports", report_headers, report_rows)

    # 15. Report–Event Links
    link_headers = ["CEID", "Event Name", "RPTID", "Report Name"]
    link_rows = [
        [3001, "ImplantStart", 3101, "RPT_ImplantStart"],
        [3001, "ImplantStart", 3103, "RPT_BeamStatus"],
        [3002, "ImplantEnd", 3102, "RPT_ImplantEnd"],
        [3002, "ImplantEnd", 3114, "RPT_DoseProfile"],
        [3003, "ImplantAbort", 3111, "RPT_ProcessAbort"],
        [3005, "BeamTuneComplete", 3103, "RPT_BeamStatus"],
        [3007, "SourceIgnition", 3107, "RPT_SourceStatus"],
        [3009, "SourceFault", 3107, "RPT_SourceStatus"],
        [3009, "SourceFault", 3106, "RPT_AlarmEvent"],
        [3010, "WaferLoaded_EndStation", 3104, "RPT_WaferLoad"],
        [3011, "WaferUnloaded_EndStation", 3105, "RPT_WaferUnload"],
        [3030, "ControlStateChange", 3110, "RPT_ControlStateChange"],
        [3031, "RecipeLoaded", 3113, "RPT_RecipeManagement"],
        [3033, "MaintenanceModeEntered", 3112, "RPT_MaintenanceEvent"],
        [3035, "HV_Enabled", 3109, "RPT_HVStatus"],
        [3037, "AlarmSet", 3106, "RPT_AlarmEvent"],
    ]
    create_sheet(wb, "Report-Event Links", link_headers, link_rows)

    # Alarm Headers
    alarm_headers = ["Alarm ID", "Name", "Code", "Linked SVID", "Description"]

    # 16. Alarms — Critical
    critical_alarm_rows = [
        [4001, "Source_ArcFault", "0x01", "1051,1052", "Arc current or voltage outside safe operating range — source off"],
        [4002, "Source_GasLeak", "0x02", "1064", "Toxic gas leak detected near ion source (AsH₃ or PH₃)"],
        [4003, "Beam_EnergyContamination", "0x03", "2064", "Energy contamination exceeds 1% — incorrect dopant depth"],
        [4004, "Beam_OverCurrent", "0x04", "1103", "Beam current exceeds safe limit for wafer damage"],
        [4005, "EndStn_DoseOverrun", "0x05", "1156,1157", "Delivered dose exceeds target by > 5% — wafer damage risk"],
        [4006, "EndStn_WaferOverheat", "0x06", "1159", "Wafer temperature exceeds 120°C — photoresist degradation"],
        [4007, "EndStn_WaferDrop", "0x07", "1160", "Wafer detected off-platen during transfer"],
        [4008, "HV_ArcOver", "0x08", "1301", "HV arc-over detected in acceleration column — HV tripped"],
        [4009, "HV_InterlockFail", "0x09", "1307", "HV interlock not satisfied but beam requested"],
        [4010, "Vac_BeamlineLeak", "0x0A", "1251", "Beamline pressure rise rate exceeds acceptable limit"],
        [4011, "Vac_EndStationLeak", "0x0B", "1252", "End station pressure rise rate exceeds acceptable limit"],
        [4012, "Robot_CollisionFault", "0x0C", "1201", "Robot collision or obstruction detected during transfer"],
        [4013, "Robot_WaferDrop", "0x0D", "1201,1202", "Wafer drop sensor activated during transfer"],
        [4014, "EMO_Activated", "0x0E", "1007", "Emergency Master Off pressed — all systems halted"],
        [4015, "Radiation_Exceeded", "0x0F", "1009", "X-ray radiation level exceeds personnel safety threshold"],
        [4016, "PowerFault_MainAC", "0x10", "1310", "Main AC power phase loss or out of tolerance"],
    ]
    create_sheet(wb, "Alarms - Critical", alarm_headers, critical_alarm_rows)

    # 17. Alarms — Warning
    warning_alarm_rows = [
        [4051, "Source_FilamentDegraded", "0x41", "1054,1061", "Filament current > 150 A to maintain arc — nearing end of life"],
        [4052, "Source_GasFlowDeviation", "0x42", "1056,1057", "Source gas flow deviates > 5% from setpoint for > 10 s"],
        [4053, "Beam_CurrentInstability", "0x43", "1103", "Beam current fluctuating > 5% during implant"],
        [4054, "Beam_ParallelismDrift", "0x44", "1114", "Beam parallelism exceeds 0.5° specification"],
        [4055, "EndStn_DoseDeviation", "0x45", "1156,1157", "Dose tracking deviates > 2% from expected profile"],
        [4056, "EndStn_TiltAngleFault", "0x46", "1151,1152", "Platen tilt angle deviates > 0.1° from setpoint"],
        [4057, "EndStn_TwistAngleFault", "0x47", "1153,1154", "Platen twist angle deviates > 0.2° from setpoint"],
        [4058, "EndStn_CoolantFlowLow", "0x48", "1162", "Backside He coolant flow below minimum"],
        [4059, "Vac_CryoPumpWarm", "0x49", "1258", "Cryopump cold head temp > 18 K — performance degraded"],
        [4060, "Vac_TurboPumpSlowdown", "0x4A", "1254,1256", "Turbopump speed decreasing — bearing wear"],
        [4061, "WH_RobotSlowResponse", "0x4B", "1201", "Robot move time exceeds expected duration by > 20%"],
        [4062, "PM_SourceDue", "0x4C", "1013", "Ion source PM counter at 80% of target"],
        [4063, "PM_BeamlineDue", "0x4D", "1014", "Beamline PM counter at 80% of target"],
        [4064, "WaferCount_Mismatch", "0x4E", "1006", "Equipment wafer count does not match host tracking"],
        [4065, "DoseUniformity_OutOfSpec", "0x4F", "2052,2012", "Post-implant dose uniformity exceeds recipe limit"],
    ]
    create_sheet(wb, "Alarms - Warning", alarm_headers, warning_alarm_rows)

    # 18. Remote Commands
    rcmd_headers = ["RCMD", "Description", "Parameters", "Resulting Event"]
    rcmd_rows = [
        ["START_IMPLANT", "Begin implant run on specified wafer with loaded recipe", "RECIPE_ID, WAFER_ID, SLOT_NUMBER, LOAD_PORT", "ImplantStart"],
        ["ABORT_IMPLANT", "Immediately abort active implant, shut beam off", "—", "ImplantAbort"],
        ["PAUSE_IMPLANT", "Pause implant; beam shutter closes, dose accumulation halted", "—", "None (SV update)"],
        ["RESUME_IMPLANT", "Resume paused implant", "—", "None (SV update)"],
        ["TUNE_BEAM", "Initiate beam tuning sequence for current recipe species/energy", "SPECIES, ENERGY_KEV", "BeamTuneStart"],
        ["SOURCE_ON", "Ignite ion source with specified gas species", "GAS_SPECIES", "SourceIgnition"],
        ["SOURCE_OFF", "Shut down ion source plasma", "—", "SourceExtinguish"],
        ["HV_ON", "Enable high-voltage acceleration", "TERMINAL_KV", "HV_Enabled"],
        ["HV_OFF", "Disable high-voltage acceleration", "—", "HV_Disabled"],
        ["LOAD_RECIPE", "Load recipe from library to active slot", "RECIPE_ID", "RecipeLoaded"],
        ["DELETE_RECIPE", "Delete recipe from controller storage", "RECIPE_ID", "RecipeDeleted"],
        ["ENTER_MAINTENANCE", "Enter maintenance mode; inhibit all processing", "—", "MaintenanceModeEntered"],
        ["EXIT_MAINTENANCE", "Exit maintenance mode; return to normal operation", "—", "MaintenanceModeExited"],
        ["RESET_FAULT", "Clear fault condition and return to IDLE", "FAULT_CODE", "None (SV update)"],
        ["SET_CLOCK", "Sync equipment clock with host", "TIMESTAMP", "EquipmentClockSync"],
        ["RESET_PM_COUNTER", "Reset PM counter for specified subsystem", "SUBSYSTEM_ID", "None"],
        ["PUMP_DOWN_LOADLOCK", "Pump down the load lock to transfer vacuum", "—", "LoadLockPumpDown"],
        ["VENT_LOADLOCK", "Vent load lock to atmosphere", "—", "LoadLockVent"],
    ]
    create_sheet(wb, "Remote Commands", rcmd_headers, rcmd_rows)

    # 19. State Definitions
    state_headers = ["State", "Applies To", "Description"]
    state_rows = [
        ["IDLE", "Equipment", "Power on, initialised, no active process, waiting for instructions"],
        ["SETTING_UP", "Equipment", "Recipe loaded, beam tuning in progress, source warming up"],
        ["BEAM_READY", "Equipment", "Source on, beam tuned and stable, ready to accept wafer"],
        ["LOADING", "End Station", "Wafer being transferred from load lock to platen"],
        ["ALIGNING", "End Station", "Platen moving to recipe tilt/twist angles"],
        ["IMPLANTING", "Equipment", "Beam on, scanning wafer, dose accumulating"],
        ["UNLOADING", "End Station", "Wafer being transferred from platen back to load lock"],
        ["COMPLETED", "Equipment", "Wafer implant complete, results available, wafer ready for unload"],
        ["FAULT", "Equipment", "Critical alarm has stopped processing; manual intervention required"],
        ["MAINTENANCE", "Equipment", "Maintenance mode active; all processing inhibited"],
    ]
    create_sheet(wb, "State Definitions", state_headers, state_rows)

    # 20. State Transitions
    trans_headers = ["From State", "To State", "Trigger", "Manual/Auto"]
    trans_rows = [
        ["IDLE", "SETTING_UP", "START_IMPLANT RCMD or TUNE_BEAM RCMD", "Manual (Host)"],
        ["SETTING_UP", "BEAM_READY", "BeamTuneComplete event — beam stable and within specification", "Auto"],
        ["BEAM_READY", "LOADING", "Wafer handler begins transfer to end station", "Auto"],
        ["LOADING", "ALIGNING", "WaferLoaded_EndStation event — wafer on platen", "Auto"],
        ["ALIGNING", "IMPLANTING", "Tilt/twist at setpoint, clamp engaged, beam shutter opens", "Auto"],
        ["IMPLANTING", "COMPLETED", "ImplantEnd event — target dose reached", "Auto"],
        ["COMPLETED", "UNLOADING", "Robot begins wafer removal from platen", "Auto"],
        ["UNLOADING", "BEAM_READY", "WaferUnloaded_EndStation event — platen empty, more wafers", "Auto"],
        ["UNLOADING", "IDLE", "WaferUnloaded & no more wafers in lot; source/HV shut down", "Auto"],
        ["IMPLANTING", "FAULT", "Any critical alarm (4001–4016)", "Auto"],
        ["SETTING_UP", "FAULT", "Source ignition failure or HV arc-over", "Auto"],
        ["BEAM_READY", "FAULT", "Beam lost or vacuum fault", "Auto"],
        ["FAULT", "IDLE", "RESET_FAULT RCMD after condition cleared", "Manual (Host)"],
        ["ANY", "MAINTENANCE", "ENTER_MAINTENANCE RCMD (from IDLE only)", "Manual (Host)"],
        ["MAINTENANCE", "IDLE", "EXIT_MAINTENANCE RCMD", "Manual (Host)"],
    ]
    create_sheet(wb, "State Transitions", trans_headers, trans_rows)

    # Message Headers
    msg_headers = ["Stream/Function", "Message Name", "Direction", "Purpose"]

    # 21. Messages H→E
    h2e_rows = [
        ["S1F1", "Are You There Request", "H→E", "Connectivity check; equipment replies S1F2"],
        ["S1F3", "Selected Equipment Status Request", "H→E", "Request values for a list of SVIDs"],
        ["S1F5", "Formatted Status Request", "H→E", "Request formatted status (StatusID-based)"],
        ["S1F13", "Establish Communications Request", "H→E", "Initiate SECS/GEM session establishment"],
        ["S2F17", "Date and Time Request", "H→E", "Request equipment current clock value"],
        ["S2F23", "Trace Initialize Send", "H→E", "Configure periodic variable trace collection"],
        ["S2F33", "Define Report", "H→E", "Create or redefine a report (RPTID + VID list)"],
        ["S2F35", "Link Event Report", "H→E", "Link a CEID to one or more RPTIDs"],
        ["S2F37", "Enable/Disable Event Report", "H→E", "Enable or disable event reporting for CEIDs"],
        ["S2F41", "Host Command Send", "H→E", "Execute a Remote Command (RCMD)"],
        ["S2F45", "Define Variable Limit Attributes", "H→E", "Set upper/lower limits for a variable"],
        ["S2F47", "Variable Limit Attribute Request", "H→E", "Query current limit settings for a variable"],
        ["S5F3", "Enable/Disable Alarm Send", "H→E", "Enable or disable specific alarm IDs"],
        ["S5F5", "List Alarms Request", "H→E", "Request list of all defined alarms and states"],
        ["S7F3", "Process Program Send", "H→E", "Upload a process recipe to equipment"],
        ["S7F5", "Process Program Request", "H→E", "Download a process recipe from equipment"],
        ["S7F17", "Delete Process Program Send", "H→E", "Delete a named recipe from equipment storage"],
        ["S7F19", "Current EPPD Request", "H→E", "Request list of all stored recipe names"],
        ["S7F25", "Enhanced Upload Process Program Send", "H→E", "Upload recipe with extended parameters (E40)"],
        ["S10F1", "Terminal Request", "H→E", "Display a text message on the operator terminal"],
    ]
    create_sheet(wb, "Messages H to E", msg_headers, h2e_rows)

    # 22. Messages E→H
    e2h_rows = [
        ["S1F2", "On-Line Data", "E→H", "Reply to S1F1 — confirms equipment is communicating"],
        ["S1F4", "Selected Equipment Status Data", "E→H", "Returns SVID values requested by S1F3"],
        ["S1F14", "Establish Communications Acknowledge", "E→H", "Acknowledges S1F13 session establishment"],
        ["S2F18", "Date and Time Data", "E→H", "Returns current equipment clock value"],
        ["S2F42", "Host Command Acknowledge", "E→H", "Acknowledge Remote Command; HCACK code in body"],
        ["S2F46", "Variable Limit Attribute Acknowledge", "E→H", "Confirm limit attribute change"],
        ["S5F1", "Alarm Report Send", "E→H", "Report an alarm set or cleared condition"],
        ["S5F6", "List Alarms Data", "E→H", "Return list of all alarms in response to S5F5"],
        ["S6F1", "Trace Data Send", "E→H", "Periodic trace data for variables set up by S2F23"],
        ["S6F11", "Event Report Send", "E→H", "Report a collection event with linked variable data"],
        ["S6F13", "Annotated Event Report Send", "E→H", "Event report with variable names included"],
        ["S7F4", "Process Program Acknowledge", "E→H", "Acknowledge receipt of uploaded recipe"],
        ["S7F6", "Process Program Data", "E→H", "Return recipe data in response to S7F5"],
        ["S7F20", "Current EPPD Data", "E→H", "Return list of all stored recipe names"],
        ["S9F1", "Unrecognised Device ID", "E→H", "Error: S/F sent with unrecognised device ID"],
        ["S9F3", "Unrecognised Stream Type", "E→H", "Error: unrecognised SECS-II stream number"],
        ["S9F5", "Unrecognised Function Type", "E→H", "Error: unrecognised SECS-II function number"],
        ["S9F7", "Illegal Data", "E→H", "Error: data in message body was malformed"],
        ["S9F9", "Transaction Timer Timeout", "E→H", "Error: reply to primary message not sent in time"],
        ["S9F11", "Data Too Long", "E→H", "Error: message body exceeded maximum allowed length"],
    ]
    create_sheet(wb, "Messages E to H", msg_headers, e2h_rows)

    # 23. GEM Compliance
    comp_headers = ["E30 Section", "Capability", "Type", "S/F Pairs", "Status"]
    comp_rows = [
        ["6.1", "Establish Communications", "Required", "S1F13/F14", "Compliant"],
        ["6.2", "Process Program Management", "Required", "S7F3/F4, S7F5/F6", "Compliant"],
        ["6.3", "Equipment Self-Description", "Required", "S1F1/F2, S1F3/F4", "Compliant"],
        ["6.4", "Control", "Required", "S2F41/F42", "Compliant"],
        ["6.5", "Alarm Management", "Required", "S5F1–F8", "Compliant"],
        ["6.6", "Event Notification", "Required", "S6F11/F12", "Compliant"],
        ["6.7", "Online Identification", "Required", "S1F1/F2", "Compliant"],
        ["6.8", "Error Messages", "Required", "S9F1–F11", "Compliant"],
        ["7.1", "Variable Data Collection", "Optional", "S2F23, S6F1–F4", "Implemented"],
        ["7.2", "Trace Data Collection", "Optional", "S2F23, S6F1", "Implemented"],
        ["7.3", "Limits Monitoring", "Optional", "S2F45–F48", "Implemented"],
        ["7.4", "Status Data Collection", "Optional", "S1F3/F4, S1F5/F6", "Implemented"],
        ["7.5", "Recipe Management Enhanced", "Optional", "S7F17–F26", "Implemented"],
        ["7.6", "Spooling", "Optional", "S2F43/F44", "Implemented"],
        ["7.7", "Clock", "Optional", "S2F17/F18", "Implemented"],
        ["7.8", "Equipment Terminal Services", "Optional", "S10F1–F4", "Implemented"],
        ["7.9", "Message Acknowledgement Spooling", "Optional", "—", "Not Implemented"],
        ["7.10", "Dynamic Event Configuration", "Optional", "S2F33/F35/F37", "Implemented"],
    ]
    create_sheet(wb, "GEM Compliance", comp_headers, comp_rows)

    # 24. Variable Cross-Reference
    xref_headers = ["Variable Name", "VID", "Type", "Chapter"]
    
    # Collect programmatically from SV & DV sheets
    vars_list = []
    
    # Global SV
    for r in sv_global_rows:
        vars_list.append([r[1], r[0], "SV", "4"])
    # Source SV
    for r in sv_source_rows:
        vars_list.append([r[1], r[0], "SV", "4"])
    # Beamline SV
    for r in sv_beamline_rows:
        vars_list.append([r[1], r[0], "SV", "4"])
    # End Station SV
    for r in sv_endstation_rows:
        vars_list.append([r[1], r[0], "SV", "4"])
    # Handler SV
    for r in sv_handler_rows:
        vars_list.append([r[1], r[0], "SV", "4"])
    # Vacuum SV
    for r in sv_vacuum_rows:
        vars_list.append([r[1], r[0], "SV", "4"])
    # HV SV
    for r in sv_hv_rows:
        vars_list.append([r[1], r[0], "SV", "4"])
    # Recipe DV
    for r in dv_recipe_rows:
        vars_list.append([r[1], r[0], "DV", "5"])
    # Results DV
    for r in dv_results_rows:
        vars_list.append([r[1], r[0], "DV", "5"])
        
    # Sort alphabetically by name
    vars_list.sort(key=lambda x: x[0].lower())
    
    create_sheet(wb, "Variable Cross-Reference", xref_headers, vars_list)

    # Save
    wb.save(output_path)
    print(f"Excel workbook successfully written to: {output_path}")

if __name__ == "__main__":
    generate_workbook(r"E:\Github\EquipmentAutomationPlatforms\IMP_HE2000_Variable_Tables.xlsx")
